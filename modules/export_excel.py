import os


def export_loss_time_record(header, produksi, catatan, manpower, absen, inhouse_claim) -> str:
    """Export loss time record ke Excel sesuai format referensi MKM. Returns filepath."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter as gcl
    except ImportError:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Loss Time Record"

    # ── Borders ───────────────────────────────────────────────────────────────
    _t  = Side(style="thin",   color="AAAAAA")
    _m  = Side(style="medium", color="000000")
    BD     = Border(left=_t, right=_t, top=_t, bottom=_t)
    BD_MED = Border(left=_m, right=_m, top=_m, bottom=_m)

    # ── Fills ─────────────────────────────────────────────────────────────────
    F_WHITE = PatternFill("solid", fgColor="FFFFFF")
    F_TEAL  = PatternFill("solid", fgColor="00B0F0")
    F_GROUP = PatternFill("solid", fgColor="1F7391")
    F_GRAY  = PatternFill("solid", fgColor="F2F2F2")
    F_TOTAL = PatternFill("solid", fgColor="D9D9D9")
    F_SHOP  = PatternFill("solid", fgColor="D9E1F2")

    # ── Fonts ─────────────────────────────────────────────────────────────────
    FN_LOGO  = Font(name="Arial", size=18, bold=True,  color="CC0000")
    FN_TITLE = Font(name="Arial", size=13, bold=True,  color="000000")
    FN_COMP  = Font(name="Arial", size=8,              color="595959")
    FN_SHOP  = Font(name="Arial", size=10, bold=True,  color="000000")
    FN_SIGN_L = Font(name="Arial", size=9, bold=True,  color="000000")
    FN_SIGN_V = Font(name="Arial", size=9,             color="000000")
    FN_GRP   = Font(name="Arial", size=9,  bold=True,  color="FFFFFF")
    FN_HDR   = Font(name="Arial", size=9,  bold=True,  color="FFFFFF")
    FN_DATA  = Font(name="Arial", size=9,              color="000000")
    FN_TOT   = Font(name="Arial", size=9,  bold=True,  color="000000")

    # ── Alignments ────────────────────────────────────────────────────────────
    ALC  = Alignment(horizontal="center", vertical="center")
    ALL  = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    ALCW = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # ── Column layout ─────────────────────────────────────────────────────────
    # 1=NO | 2=DATE | 3=SHIFT | 4=OP Number
    # 5-7=PROBLEM (3 cols, merged) | 8=Category
    # 9-12=COUNTER MEASURE (4 cols, merged) | 13=Remarks
    # 14=Start | 15=Finish | 16=HOUR
    # 17=M/C | 18=LINE STOP(HOUR) | 19=M/C STOP(HOUR) | 20=MAN | 21=MAN HOUR
    NCOLS  = 21
    C_NO   =  1
    C_DT   =  2
    C_SH   =  3
    C_OP   =  4   # OP Number  (ra_number)
    C_PRS  =  5   # PROBLEM start
    C_PRE  =  7   # PROBLEM end  (spans 5-7)
    C_CAT  =  8   # Category
    C_CMS  =  9   # COUNTER MEASURE start
    C_CME  = 12   # COUNTER MEASURE end  (spans 9-12)
    C_REM  = 13   # Remarks
    C_ST   = 14   # Start  (TIME)
    C_FN   = 15   # Finish (TIME)
    C_HR   = 16   # HOUR   (TIME)
    C_MC   = 17   # M/C    (LOSS TIME)
    C_LS   = 18   # LINE STOP
    C_MS   = 19   # M/C STOP
    C_MAN  = 20   # MAN
    C_MH   = 21   # MAN HOUR

    DATA_ROW = 9

    # ── Column widths ─────────────────────────────────────────────────────────
    col_w = {
        C_NO: 5,   C_DT: 11,  C_SH: 9,   C_OP: 14,
        C_PRS: 16, 6: 16, C_PRE: 16,
        C_CAT: 10,
        C_CMS: 15, 10: 15, 11: 15, C_CME: 15,
        C_REM: 11,
        C_ST: 7,  C_FN: 7,  C_HR: 7,
        C_MC: 7,  C_LS: 10, C_MS: 10, C_MAN: 7, C_MH: 8,
    }
    for ci, w in col_w.items():
        ws.column_dimensions[gcl(ci)].width = w

    # ── Helpers ───────────────────────────────────────────────────────────────
    def mrg(r1, c1, r2, c2):
        if r1 == r2 and c1 == c2:
            return
        ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)

    def wc(r, col, val="", fn=None, fl=None, al=None, bd=None):
        cl = ws.cell(row=r, column=col, value=val)
        if fn: cl.font      = fn
        if fl: cl.fill      = fl
        if al: cl.alignment = al
        if bd: cl.border    = bd
        return cl

    def fill_range(r, c1, c2, fl=F_WHITE, fn=FN_DATA, bd=BD):
        for c in range(c1, c2 + 1):
            cl = ws.cell(r, c)
            if fl: cl.fill   = fl
            if fn: cl.font   = fn
            if bd: cl.border = bd

    # ── Row heights ───────────────────────────────────────────────────────────
    for r, h in {1: 16, 2: 30, 3: 14, 4: 22, 5: 8, 6: 8, 7: 18, 8: 32}.items():
        ws.row_dimensions[r].height = h

    # =========================================================================
    # HEADER AREA  (rows 1–6)
    # =========================================================================

    # Row 1 — Approve / Checked / Prepared labels
    mrg(1, C_MC, 1, C_LS)
    wc(1, C_MC,  "Approve",  FN_SIGN_L, F_WHITE, ALC, BD)
    wc(1, C_MS,  "Checked",  FN_SIGN_L, F_WHITE, ALC, BD)
    mrg(1, C_MAN, 1, C_MH)
    wc(1, C_MAN, "Prepared", FN_SIGN_L, F_WHITE, ALC, BD)

    # Row 2 — Logo | Title | Names
    mrg(2, C_NO, 2, C_SH)
    wc(2, C_NO, "MKM", FN_LOGO, F_WHITE, ALC)

    mrg(2, C_OP, 2, C_REM)
    wc(2, C_OP,
       f"LOSS TIME RECORD   {header.get('section', '').upper()}",
       FN_TITLE, F_WHITE, ALC)

    mrg(2, C_MC, 2, C_LS)
    wc(2, C_MC,  header.get("approved_by", "—"), FN_SIGN_V, F_WHITE, ALC, BD)
    wc(2, C_MS,  header.get("checked_by",  "—"), FN_SIGN_V, F_WHITE, ALC, BD)
    mrg(2, C_MAN, 2, C_MH)
    wc(2, C_MAN, header.get("coordinator", "—"), FN_SIGN_V, F_WHITE, ALC, BD)

    # Row 3 — Company name
    mrg(3, C_NO, 3, C_REM)
    wc(3, C_NO, "PT. MITSUBISHI KRAMAYUDHA MOTORS & MFG", FN_COMP, F_WHITE, ALC)

    # Row 4 — SHOP + date/shift
    mrg(4, C_DT, 4, C_FN)
    wc(4, C_DT,
       f"SHOP  :  {header.get('section', '')}",
       FN_SHOP, F_SHOP, ALC, BD_MED)

    mrg(4, C_HR, 4, C_MH)
    wc(4, C_HR,
       f"Tanggal : {header.get('date', '')}     Shift : {header.get('shift', '')}",
       FN_DATA, F_WHITE, ALC)

    # =========================================================================
    # TABLE HEADERS  (rows 7–8)
    # =========================================================================

    # Row 7 — Group headers
    fill_range(7, C_NO, NCOLS, fl=F_TEAL, fn=FN_HDR, bd=BD)
    mrg(7, C_ST, 7, C_HR)
    wc(7, C_ST,  "TIME",      FN_GRP, F_GROUP, ALC, BD)
    mrg(7, C_MC, 7, C_MH)
    wc(7, C_MC, "LOSS TIME",  FN_GRP, F_GROUP, ALC, BD)

    # Row 8 — Column labels
    fill_range(8, C_NO, NCOLS, fl=F_TEAL, fn=FN_HDR, bd=BD)
    mrg(8, C_PRS, 8, C_PRE)
    mrg(8, C_CMS, 8, C_CME)

    lbl8 = {
        C_NO:  "NO.",
        C_DT:  "DATE",
        C_SH:  "SHIFT",
        C_OP:  "OP Number",
        C_PRS: "PROBLEM",
        C_CAT: "Category",
        C_CMS: "COUNTER MEASURE",
        C_REM: "Remarks",
        C_ST:  "Start",
        C_FN:  "Finish",
        C_HR:  "HOUR",
        C_MC:  "M/C",
        C_LS:  "LINE STOP\n(HOUR)",
        C_MS:  "M/C STOP\n(HOUR)",
        C_MAN: "MAN",
        C_MH:  "MAN\nHOUR",
    }
    for ci, text in lbl8.items():
        wc(8, ci, text, FN_HDR, F_TEAL, ALCW, BD)

    # =========================================================================
    # DATA ROWS
    # =========================================================================

    for i, item in enumerate(catatan, 1):
        r = DATA_ROW + i - 1
        ws.row_dimensions[r].height = 20

        fl = F_WHITE if i % 2 == 1 else F_GRAY
        fill_range(r, C_NO, NCOLS, fl=fl, fn=FN_DATA, bd=BD)

        mrg(r, C_PRS, r, C_PRE)
        mrg(r, C_CMS, r, C_CME)

        wc(r, C_NO,  i,                                       al=ALC)
        wc(r, C_DT,  header.get("date",  ""),                 al=ALC)
        wc(r, C_SH,  header.get("shift", ""),                 al=ALC)
        wc(r, C_OP,  item.get("ra_number", "") or "",         al=ALC)
        wc(r, C_PRS, item.get("description", "") or "",       al=ALL)
        wc(r, C_CAT, item.get("category", "") or "",          al=ALC)
        wc(r, C_CMS, item.get("corrective_action", "") or "", al=ALL)
        wc(r, C_REM, item.get("cause", "") or "",             al=ALL)
        wc(r, C_ST,  item.get("start_time", ""),              al=ALC)
        wc(r, C_FN,  item.get("end_time",   ""),              al=ALC)

        down = float(item.get("down_time", 0) or 0)
        loss = float(item.get("loss_time", 0) or 0)

        for ci, val in [(C_HR, down), (C_MC, 0.0), (C_LS, loss),
                        (C_MS, 0.0), (C_MAN, 0.0), (C_MH, 0.0)]:
            cl = wc(r, ci, val, al=ALC)
            cl.number_format = "0.00"

    # =========================================================================
    # TOTAL ROW
    # =========================================================================

    if catatan:
        r_tot = DATA_ROW + len(catatan)
        ws.row_dimensions[r_tot].height = 20

        fill_range(r_tot, C_NO, NCOLS, fl=F_TOTAL, fn=FN_TOT, bd=BD)
        mrg(r_tot, C_NO, r_tot, C_REM)
        wc(r_tot, C_NO, "TOTAL", FN_TOT, F_TOTAL, ALC, BD)

        t_down = sum(float(x.get("down_time") or 0) for x in catatan)
        t_loss = sum(float(x.get("loss_time") or 0) for x in catatan)

        for ci, val in [(C_HR, t_down), (C_MC, 0.0), (C_LS, t_loss),
                        (C_MS, 0.0), (C_MAN, 0.0), (C_MH, 0.0)]:
            cl = wc(r_tot, ci, val, FN_TOT, F_TOTAL, ALC, BD)
            cl.number_format = "0.00"

    ws.freeze_panes = ws.cell(DATA_ROW, C_NO)

    # =========================================================================
    # SAVE
    # =========================================================================
    downloads   = os.path.join(os.path.expanduser("~"), "Downloads")
    date_str    = str(header.get("date", "")).replace("-", "")
    section_str = str(header.get("section", "unknown")).replace(" ", "_")
    filename    = f"LTR_{section_str}_{date_str}.xlsx"
    filepath    = os.path.join(downloads, filename)
    wb.save(filepath)
    return filepath


# =============================================================================
# EXPORT INHOUSE CLAIM NG & PART PENDING
# =============================================================================

def export_inhouse_ng_pending(ng_rows, pending_rows, section_name, period_str) -> str:
    """
    Export data NG & Part Pending ke satu sheet Excel (digabung, dengan kolom Status).

    ng_rows      : list of 13-element lists  — kolom terakhir = Status (NG/dll)
    pending_rows : list of 12-element lists  — tanpa Status; otomatis ditambah "PENDING"
    section_name : nama section untuk judul
    period_str   : string periode, misal "MEI 2026"
    Returns filepath.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter as gcl
    except ImportError:
        raise ImportError("openpyxl tidak terinstall. Jalankan: pip install openpyxl")

    # ── Styles ────────────────────────────────────────────────────────────────
    _thin  = Side(style="thin",   color="000000")
    _thick = Side(style="medium", color="000000")
    BD     = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    BD_TTL = Border(left=_thick, right=_thick, top=_thick, bottom=_thick)

    F_YEL = PatternFill("solid", fgColor="FFFF00")
    F_WHT = PatternFill("solid", fgColor="FFFFFF")

    FN_TTL = Font(name="Arial", size=11, bold=True, color="000000")
    FN_HDR = Font(name="Arial", size=10, bold=True, color="000000")
    FN_DAT = Font(name="Arial", size=9,             color="000000")
    FN_NG  = Font(name="Arial", size=9,  bold=True, color="FF0000")
    FN_PND = Font(name="Arial", size=9,  bold=True, color="FF7700")

    ALC = Alignment(horizontal="center", vertical="center")
    ALL = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    HEADERS = [
        "No", "Tanggal", "Model", "OP.No/St.", "Item", "Qty", "Satuan",
        "Penyebab", "Tindakan", "Faktor", "Stop (Hr)", "Lost (Hr)", "Status",
    ]
    NCOLS = len(HEADERS)
    COL_W = [5, 12, 10, 10, 22, 6, 7, 20, 24, 12, 10, 10, 8]
    CTR   = {1, 2, 3, 4, 6, 7, 10, 11, 12, 13}   # center-aligned columns (1-based)

    # ── Gabungkan NG + Pending jadi satu list ─────────────────────────────────
    combined = []
    for row in ng_rows:
        combined.append(list(row))                     # sudah ada Status di index -1
    for row in pending_rows:
        r = list(row)
        if len(r) < NCOLS:
            r.append("PENDING")                        # tambahkan Status
        combined.append(r)

    # Renumber kolom No (index 0) sesuai urutan gabungan
    for seq, row in enumerate(combined, 1):
        row[0] = seq

    # ── Build sheet ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inhouse Claim"

    for i, w in enumerate(COL_W, 1):
        ws.column_dimensions[gcl(i)].width = w

    # Row 1: merged title
    ws.row_dimensions[1].height = 22
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=NCOLS)
    title_val = f"DATA INHOUSE CLAIM {section_name.upper()} {period_str.upper()}"
    cl = ws.cell(row=1, column=1, value=title_val)
    cl.font = FN_TTL; cl.fill = F_YEL; cl.alignment = ALC; cl.border = BD_TTL

    # Row 2: column headers
    ws.row_dimensions[2].height = 20
    for ci, h in enumerate(HEADERS, 1):
        cl = ws.cell(row=2, column=ci, value=h)
        cl.font = FN_HDR; cl.fill = F_YEL; cl.alignment = ALC; cl.border = BD

    # Data rows
    for i, row in enumerate(combined, 1):
        r = i + 2
        ws.row_dimensions[r].height = 18
        for ci in range(1, NCOLS + 1):
            val = row[ci - 1] if ci - 1 < len(row) else ""
            cl  = ws.cell(row=r, column=ci, value=val)
            cl.fill      = F_WHT
            cl.border    = BD
            cl.alignment = ALC if ci in CTR else ALL

            if ci == NCOLS:                            # kolom Status
                s = str(val) if val else ""
                cl.font = FN_NG if s == "NG" else FN_PND if s == "PENDING" else FN_DAT
            else:
                cl.font = FN_DAT

            if ci in (11, 12):
                cl.number_format = "0.0000"

    ws.freeze_panes = ws.cell(3, 1)

    # ── Save ──────────────────────────────────────────────────────────────────
    downloads = os.path.join(os.path.expanduser("~"), "Downloads")
    safe_sec  = section_name.replace(" ", "_") if section_name else "ALL"
    safe_per  = period_str.replace("/", "-").replace(" ", "_")
    filename  = f"IC_NG_Pending_{safe_sec}_{safe_per}.xlsx"
    filepath  = os.path.join(downloads, filename)
    wb.save(filepath)
    return filepath
