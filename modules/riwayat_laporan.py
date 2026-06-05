import os

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QFrame, QComboBox, QPushButton, QDateEdit,
    QScrollArea, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QDialog, QAbstractItemView,
    QTabWidget, QLineEdit, QStyledItemDelegate, QApplication,
)
from PySide6.QtCore import Qt, QDate, QThread, Signal, QSize, QRect
from PySide6.QtGui import QColor, QTextDocument

from modules.db_auth import get_connection, release_connection
from modules.db_laporan import (
    get_all_sections, get_all_shifts, get_all_category_names,
    get_detail_laporan, hapus_laporan, get_monthly_productivity,
    get_production_volume, get_riwayat_laporan,
)
from modules.export_excel import export_loss_time_record


# Shared styles

# (prev: #1e1e1e bg, #303030 border, #252525 card, #969696 text-sec)

_DATE_STYLE = """
    QDateEdit {
        background-color: #2a2a2a;
        border: 1px solid #2e2e2e;
        border-radius: 0px; padding-left: 8px;
        color: #f0f0f0; font-size: 11px;
    }
    QDateEdit:focus { border: 1px solid #da291c; }
    QDateEdit::drop-down {
        border: none; width: 22px;
        subcontrol-origin: padding; subcontrol-position: top right;
    }
"""

_COMBO_STYLE = """
    QComboBox {
        background-color: #2a2a2a;
        border: 1px solid #2e2e2e;
        border-radius: 0px; padding-left: 8px;
        color: #f0f0f0; font-size: 11px;
    }
    QComboBox:focus { border: 1px solid #da291c; }
    QComboBox::drop-down { border: none; width: 24px; }
    QComboBox QAbstractItemView {
        background-color: #1a1a1a; color: #f0f0f0;
        selection-background-color: #2a2a2a;
        border: 1px solid #2e2e2e;
    }
"""

_TABLE_STYLE = """
    QTableWidget {
        background-color: #1a1a1a;
        border: 1px solid #2e2e2e; border-radius: 0px;
        gridline-color: #2e2e2e;
    }
    QTableWidget::item {
        color: #f0f0f0; padding: 4px 8px;
        background-color: #222222;
        border-bottom: 1px solid #2e2e2e;
    }
    QTableWidget::item:alternate { background-color: #1e1e1e; }
    QTableWidget::item:selected { background-color: #2a2a2a; color: #ffffff; }
    QHeaderView::section {
        background-color: #111111; color: #888888;
        border: none; border-bottom: 1px solid #2e2e2e;
        border-right: 1px solid #2e2e2e;
        padding: 5px 8px; font-weight: bold; font-size: 10px;
        text-transform: uppercase; letter-spacing: 1px;
    }
"""

_TAB_STYLE = """
    QTabWidget::pane { background-color: transparent; border: none; }
    QTabBar::tab {
        background-color: #1a1a1a; color: #555555;
        padding: 8px 22px; margin-right: 2px;
        border-radius: 0px; font-size: 10px; font-weight: bold;
        letter-spacing: 1px; text-transform: uppercase;
    }
    QTabBar::tab:selected {
        background-color: #222222; color: #f0f0f0;
        border-bottom: 2px solid #da291c;
    }
    QTabBar::tab:hover:!selected {
        background-color: #222222; color: #f0f0f0;
    }
"""

_INPUT_STYLE = """
    QLineEdit {
        background-color: #2a2a2a;
        border: 1px solid #2e2e2e; border-radius: 0px;
        padding-left: 8px; color: #f0f0f0; font-size: 11px;
    }
    QLineEdit:focus { border: 1px solid #da291c; }
"""

_CARD_STYLE = "QFrame { background-color: #222222; border-radius: 0px; }"

_BTN_CARI = """
    QPushButton {
        background-color: #da291c; color: #ffffff;
        border: none; border-radius: 0px; font-size: 11px; padding: 0 12px;
        text-transform: uppercase; letter-spacing: 1px;
    }
    QPushButton:hover { background-color: #b01e0a; }
"""

_BTN_RESET = """
    QPushButton {
        background-color: #2a2a2a; color: #888888;
        border: 1px solid #3a3a3a; border-radius: 0px; font-size: 11px; padding: 0 12px;
    }
    QPushButton:hover { background-color: #333333; color: #f0f0f0; }
"""

_BTN_EXPORT = """
    QPushButton {
        background-color: #1a2a1a; color: #22863a;
        border: 1px solid #1a4a1a; border-radius: 0px; font-size: 11px; padding: 0 12px;
    }
    QPushButton:hover { background-color: #1e341e; }
"""


_HDR_LBL_D = ("color: #f0f0f0; font-size: 11px; font-weight: bold;"
              " border-left: 2px solid #da291c; padding-left: 8px;"
              " letter-spacing: 1px; text-transform: uppercase;")
_FLD_LBL_D = "color: #888888; font-size: 10px; letter-spacing: 1px;"


def _fmt(v: float) -> str:
    """Format angka dinamis: hilangkan trailing zero, max 4 desimal."""
    s = f"{v:.4f}".rstrip('0').rstrip('.')
    return s or "0"


class _WrapDelegate(QStyledItemDelegate):
    """Read-only delegate: paint multi-line text via QTextDocument, auto row height."""

    def __init__(self, table, parent=None):
        super().__init__(parent)
        self._table = table

    def paint(self, painter, option, index):
        text = index.data(Qt.DisplayRole) or ""
        if not text:
            super().paint(painter, option, index)
            return
        painter.save()
        doc = QTextDocument()
        doc.setDefaultFont(option.font)
        doc.setPlainText(text)
        doc.setTextWidth(option.rect.width() - 12)
        painter.setClipRect(option.rect)
        painter.translate(option.rect.topLeft())
        painter.translate(6, 4)
        doc.drawContents(painter)
        painter.restore()

    def sizeHint(self, option, index):
        text = index.data(Qt.DisplayRole) or ""
        if not text:
            return QSize(option.rect.width(), 30)
        fm = option.fontMetrics
        col_w = max(option.rect.width() - 12, 40)
        rect = fm.boundingRect(
            QRect(0, 0, col_w, 10000),
            Qt.TextWordWrap | Qt.AlignLeft,
            text,
        )
        return QSize(option.rect.width(), max(30, rect.height() + 8))


#  DB helpers

def _db_display_line_stop(section_id, bulan, tahun, factor=None):
    conn = get_connection()
    try:
        cur = conn.cursor()
        sec = " AND dr.section_id = %s" if section_id is not None else ""
        p = [bulan, tahun] + ([section_id] if section_id else [])

        cur.execute(f"""
            SELECT COALESCE(SUM(s.total_hours), 0)
            FROM daily_report dr
            JOIN shift s ON s.id = dr.shift_id
            WHERE EXTRACT(MONTH FROM dr.date) = %s
              AND EXTRACT(YEAR  FROM dr.date) = %s {sec}
        """, p)
        total_hour = float((cur.fetchone() or [0])[0] or 0)

        cur.execute(f"""
            SELECT COALESCE(SUM(
                CASE WHEN dp.ot_2h  THEN 2.0  ELSE 0 END +
                CASE WHEN dp.ot_3h  THEN 3.0  ELSE 0 END +
                CASE WHEN dp.ot_11h THEN 11.0 ELSE 0 END
            ), 0)
            FROM (SELECT DISTINCT ON (report_id) report_id, ot_2h, ot_3h, ot_11h
                  FROM daily_production) dp
            JOIN daily_report dr ON dr.id = dp.report_id
            WHERE EXTRACT(MONTH FROM dr.date) = %s
              AND EXTRACT(YEAR  FROM dr.date) = %s {sec}
        """, p)
        total_hour += float((cur.fetchone() or [0])[0] or 0)

        extra = ""
        p2 = list(p)
        if factor and factor != "Semua":
            extra += " AND pc.name = %s"
            p2.append(factor)

        cur.execute(f"""
            SELECT
                dr.date,
                pr.ra_number,
                pr.description,
                pr.cause,
                pr.corrective_action,
                COALESCE(pc.name, 'Others') AS factor,
                COALESCE(pr.loss_time, 0),
                COALESCE(pr.down_time, 0)
            FROM problem_record pr
            JOIN daily_report dr ON dr.id = pr.report_id
            LEFT JOIN problem_category pc ON pc.id = pr.category_id
            WHERE EXTRACT(MONTH FROM dr.date) = %s
              AND EXTRACT(YEAR  FROM dr.date) = %s {sec} {extra}
            ORDER BY dr.date ASC, pr.id ASC
        """, p2)
        rows = cur.fetchall()
        cur.close()
        return rows, total_hour
    finally:
        release_connection(conn)


# Background worker

class _RiwayatWorker(QThread):
    finished = Signal(list)
    error    = Signal(str)

    def __init__(self, section_id, shift_name, date_from, date_to):
        super().__init__()
        self.section_id = section_id
        self.shift_name = shift_name
        self.date_from  = date_from
        self.date_to    = date_to

    def run(self):
        try:
            rows = get_riwayat_laporan(
                section_id=self.section_id,
                shift_name=self.shift_name,
                date_from=self.date_from,
                date_to=self.date_to,
            )
            self.finished.emit(rows)
        except Exception as e:
            self.error.emit(str(e))


# Widget

class RiwayatLaporanWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._sections_loaded = False
        self._setup_ui()

    def showEvent(self, event):
        super().showEvent(event)
        self._load_sections_once()
        self.load_data_harian()

    def _load_sections_once(self):
        if self._sections_loaded:
            return
        try:
            sections = get_all_sections()
            # combo_section (Riwayat Harian) — tanpa item "Semua", shop wajib dipilih
            self.combo_section.blockSignals(True)
            for sid, sname in sections:
                self.combo_section.addItem(sname, sid)
            self.combo_section.blockSignals(False)
            # combo lain — dengan item "Semua"
            for combo in (self.combo_section_rekap, self.combo_prod_section):
                combo.blockSignals(True)
                for sid, sname in sections:
                    combo.addItem(sname, sid)
                combo.blockSignals(False)
            # combo volume produksi — tanpa "Semua", shop wajib dipilih
            self.combo_vol_section.blockSignals(True)
            for sid, sname in sections:
                self.combo_vol_section.addItem(sname, sid)
            self.combo_vol_section.blockSignals(False)
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", f"Gagal memuat daftar shop: {e}")
        try:
            shifts = get_all_shifts()
            self.combo_shift.blockSignals(True)
            self.combo_shift.clear()
            self.combo_shift.addItem("Semua", None)
            for s in shifts:
                self.combo_shift.addItem(s["name"])
            self.combo_shift.blockSignals(False)
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", f"Gagal memuat daftar shift: {e}")
        self._sections_loaded = True

    def _on_tab_changed(self, idx):
        if idx == 1:
            self.load_rekap_bulanan()
        elif idx == 2:
            self.load_produktivitas()
        elif idx == 3:
            pass  # Volume Produksi — tidak auto-load, user harus pilih shop dulu

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(_TAB_STYLE)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        self.tabs.addTab(self._build_tab_harian(), "Riwayat Harian")
        self.tabs.addTab(self._build_tab_rekap(), "Rekap Bulanan")
        self.tabs.addTab(self._build_tab_produktivitas(), "Produktivitas")
        self.tabs.addTab(self._build_tab_volume(),        "Volume Produksi")
        outer.addWidget(self.tabs)

    # Tab 1: Riwayat Harian

    def _build_tab_harian(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_section = QComboBox()
        self.combo_section.setMinimumHeight(30)
        self.combo_section.setMinimumWidth(155)
        self.combo_section.setStyleSheet(_COMBO_STYLE)
        fl.addWidget(self.combo_section)

        fl.addWidget(self._flabel("Shift"))
        self.combo_shift = QComboBox()
        self.combo_shift.setMinimumHeight(30)
        self.combo_shift.setMinimumWidth(110)
        self.combo_shift.setStyleSheet(_COMBO_STYLE)
        self.combo_shift.addItem("Semua", None)
        fl.addWidget(self.combo_shift)

        fl.addWidget(self._flabel("Dari"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setMinimumHeight(30)
        self.date_from.setStyleSheet(_DATE_STYLE)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        fl.addWidget(self.date_from)

        fl.addWidget(self._flabel("Sampai"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setMinimumHeight(30)
        self.date_to.setStyleSheet(_DATE_STYLE)
        self.date_to.setDate(QDate.currentDate())
        fl.addWidget(self.date_to)

        fl.addStretch()

        btn_cari = QPushButton("Cari")
        btn_cari.setMinimumSize(76, 32)
        btn_cari.setStyleSheet(_BTN_CARI)
        btn_cari.clicked.connect(self.load_data_harian)
        fl.addWidget(btn_cari)

        btn_reset_f = QPushButton("Reset Filter")
        btn_reset_f.setMinimumSize(90, 32)
        btn_reset_f.setStyleSheet(_BTN_RESET)
        btn_reset_f.clicked.connect(self._reset_filter_harian)
        fl.addWidget(btn_reset_f)

        main.addWidget(card_f)

        # Table card
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        hdr_row = QHBoxLayout()
        self.lbl_info_h = QLabel("Memuat data...")
        self.lbl_info_h.setStyleSheet("color: #969696; font-size: 10px;")
        hdr_row.addWidget(self.lbl_info_h)
        hdr_row.addStretch()
        btn_export_rah = QPushButton("Export RAH")
        btn_export_rah.setMinimumSize(100, 30)
        btn_export_rah.setStyleSheet(_BTN_EXPORT)
        btn_export_rah.clicked.connect(self._export_rah)
        hdr_row.addWidget(btn_export_rah)
        tl.addLayout(hdr_row)

        self.tabel_harian = QTableWidget()
        self.tabel_harian.setColumnCount(7)
        self.tabel_harian.setHorizontalHeaderLabels([
            "No", "Tanggal", "Shop", "Shift", "Koordinator", "Jml Masalah", "Aksi",
        ])
        self.tabel_harian.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabel_harian.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.tabel_harian.horizontalHeader().setSectionResizeMode(6, QHeaderView.Fixed)
        self.tabel_harian.horizontalHeader().setStretchLastSection(False)
        self.tabel_harian.verticalHeader().setVisible(False)
        self.tabel_harian.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_harian.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_harian.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tabel_harian.setAlternatingRowColors(True)
        self.tabel_harian.setStyleSheet(_TABLE_STYLE)
        self.tabel_harian.setColumnWidth(0, 40)
        self.tabel_harian.setColumnWidth(1, 92)
        self.tabel_harian.setColumnWidth(3, 92)
        self.tabel_harian.setColumnWidth(4, 135)
        self.tabel_harian.setColumnWidth(5, 90)
        self.tabel_harian.setColumnWidth(6, 290)
        tl.addWidget(self.tabel_harian)

        main.addWidget(card_t)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_data_harian(self):
        section_id = self.combo_section.currentData()
        shift_text = self.combo_shift.currentText()
        shift_name = None if shift_text == "Semua" else shift_text
        date_from  = self.date_from.date().toString("yyyy-MM-dd")
        date_to    = self.date_to.date().toString("yyyy-MM-dd")

        self.lbl_info_h.setText("Memuat data...")
        self.tabel_harian.setRowCount(0)

        self._worker = _RiwayatWorker(section_id, shift_name, date_from, date_to)
        self._worker.finished.connect(self._on_harian_loaded)
        self._worker.error.connect(lambda e: self.lbl_info_h.setText(f"Gagal: {e}"))
        self._worker.start()

    def _on_harian_loaded(self, rows: list):
        self._loaded_rows = rows
        self.tabel_harian.setRowCount(0)
        self.lbl_info_h.setText(f"{len(rows)} data ditemukan")

        def _item(text, align=Qt.AlignCenter):
            it = QTableWidgetItem(str(text) if text is not None else "")
            it.setTextAlignment(align)
            return it

        for i, row in enumerate(rows):
            self.tabel_harian.insertRow(i)
            self.tabel_harian.setItem(i, 0, _item(i + 1))
            self.tabel_harian.setItem(i, 1, _item(row["date"]))
            self.tabel_harian.setItem(i, 2, _item(row["section"], Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_harian.setItem(i, 3, _item(row["shift"]))
            self.tabel_harian.setItem(i, 4, _item(row["coordinator"], Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_harian.setItem(i, 5, _item(str(row["jml_masalah"])))

            report_id = row["id"]
            aksi_w = QWidget()
            aksi_w.setStyleSheet("background-color: #252525;")
            al = QHBoxLayout(aksi_w)
            al.setContentsMargins(8, 0, 8, 0)
            al.setSpacing(8)
            al.setAlignment(Qt.AlignVCenter | Qt.AlignCenter)

            btn_lihat = QPushButton("Lihat")
            btn_lihat.setMinimumWidth(48); btn_lihat.setFixedHeight(24)
            btn_lihat.setStyleSheet(
                "QPushButton { background-color: #1a2a3a; color: #4a9fd4;"
                " border: 1px solid #1a4a6a; border-radius: 0px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #1e344a; }"
            )
            btn_lihat.clicked.connect(lambda _, rid=report_id: self._lihat(rid))

            btn_edit = QPushButton("Edit")
            btn_edit.setMinimumWidth(44); btn_edit.setFixedHeight(26)
            btn_edit.setStyleSheet(
                "QPushButton { background-color: #2a2a1a; color: #b08800;"
                " border: 1px solid #4a4a1a; border-radius: 0px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #34341e; }"
            )
            btn_edit.clicked.connect(lambda _, rid=report_id: self._edit(rid))
            btn_edit.setVisible(False)  # akan diimplementasi di sprint berikutnya

            btn_hapus = QPushButton("Hapus")
            btn_hapus.setMinimumWidth(48); btn_hapus.setFixedHeight(24)
            btn_hapus.setStyleSheet(
                "QPushButton { background-color: #2a1a1a; color: #da291c;"
                " border: 1px solid #4a1a1a; border-radius: 0px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #341e1e; }"
            )
            btn_hapus.clicked.connect(lambda _, rid=report_id: self._hapus(rid))

            al.addWidget(btn_lihat)
            al.addWidget(btn_edit)
            al.addWidget(btn_hapus)
            self.tabel_harian.setCellWidget(i, 6, aksi_w)
            self.tabel_harian.setRowHeight(i, 36)

    def _reset_filter_harian(self):
        self.combo_section.setCurrentIndex(0)
        self.combo_shift.setCurrentIndex(0)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to.setDate(QDate.currentDate())
        self.load_data_harian()

    def _lihat(self, report_id):
        try:
            header, produksi, catatan, _, absen, inhouse_claim, materials = get_detail_laporan(report_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat detail: {e}")
            return
        if header is None:
            QMessageBox.warning(self, "Tidak Ditemukan", f"Laporan #{report_id} tidak ditemukan.")
            return

        # ── Shift lookup untuk Calculation Hour ───────────────────────────
        try:
            shifts     = get_all_shifts()
            shift_data = next((s for s in shifts if s["name"] == header.get("shift", "")), {})
        except Exception:
            shift_data = {}
        total_hour = shift_data.get("total_hours", 0.0)
        prep_h     = shift_data.get("preparation_min", 15.0) / 60
        other_h    = shift_data.get("sholat_min", 10.0) / 60

        process  = sum(p.get("actual_whour", 0) or 0 for p in produksi)
        linestop = sum(c.get("loss_time", 0) or 0 for c in catatan)
        absence  = sum(
            float(a["keterangan"])
            for a in absen
            if str(a.get("keterangan", "")).replace(".", "").isdigit()
        )
        balance  = total_hour - process - prep_h - other_h - linestop - absence

        # ── Dialog ────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Detail Laporan — #{report_id}")
        dlg.setMinimumSize(1200, 700)
        dlg.resize(1300, 800)
        dlg.setStyleSheet("QDialog { background-color: #1e1e1e; color: #ffffff; }")
        _screen = QApplication.primaryScreen().geometry()
        dlg.move(
            _screen.center().x() - dlg.width() // 2,
            _screen.center().y() - dlg.height() // 2,
        )

        outer_lyt = QVBoxLayout(dlg)
        outer_lyt.setContentsMargins(0, 0, 0, 0)
        outer_lyt.setSpacing(0)

        # ── Helper ────────────────────────────────────────────────────────
        def _c(val, align=Qt.AlignLeft | Qt.AlignVCenter):
            it = QTableWidgetItem(str(val) if val is not None else "")
            it.setTextAlignment(align)
            it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            return it

        def _card():
            f = QFrame()
            f.setStyleSheet(_CARD_STYLE)
            return f

        def _hdr_lbl(text):
            l = QLabel(text)
            l.setStyleSheet(_HDR_LBL_D)
            return l

        def _tbl(cols, headers, stretch_cols=(), fixed_cols=()):
            t = QTableWidget(0, cols)
            t.setHorizontalHeaderLabels(headers)
            h = t.horizontalHeader()
            h.setSectionResizeMode(QHeaderView.Stretch)
            for c in stretch_cols:
                h.setSectionResizeMode(c, QHeaderView.Stretch)
            for c, w in fixed_cols:
                h.setSectionResizeMode(c, QHeaderView.Fixed)
                t.setColumnWidth(c, w)
            t.verticalHeader().setVisible(False)
            t.setEditTriggers(QAbstractItemView.NoEditTriggers)
            t.setSelectionBehavior(QAbstractItemView.SelectRows)
            t.setAlternatingRowColors(True)
            t.setStyleSheet(_TABLE_STYLE)
            return t

        # ── HEADER BAR ────────────────────────────────────────────────────
        hdr_bar = QFrame()
        hdr_bar.setFixedHeight(70)
        hdr_bar.setStyleSheet(
            "QFrame { background-color: #1a1a1a;"
            " border-bottom: 2px solid #da291c; }"
        )
        hdr_row = QHBoxLayout(hdr_bar)
        hdr_row.setContentsMargins(16, 10, 16, 10)
        hdr_row.setSpacing(0)

        _days = ["SENIN", "SELASA", "RABU", "KAMIS", "JUMAT", "SABTU", "MINGGU"]
        try:
            from datetime import date as _dt
            _d    = _dt.fromisoformat(header.get("date", ""))
            _hari = _days[_d.weekday()]
        except Exception:
            _hari = "—"

        overtime = header.get("overtime", "-")

        def _make_field(label, value):
            w = QWidget()
            w.setStyleSheet("background: transparent;")
            v = QVBoxLayout(w)
            v.setContentsMargins(0, 0, 0, 0)
            v.setSpacing(2)
            lbl = QLabel(label)
            lbl.setStyleSheet(
                "color:#666666; font-size:9px; letter-spacing:1px;"
                " background:transparent;"
            )
            val = QLabel(str(value))
            val.setStyleSheet(
                "color:#f0f0f0; font-size:12px; font-weight:bold;"
                " background:transparent;"
            )
            v.addWidget(lbl)
            v.addWidget(val)
            return w

        def _vsep_h():
            sep = QFrame()
            sep.setFrameShape(QFrame.VLine)
            sep.setFixedWidth(1)
            sep.setStyleSheet("background:#333333; border:none;")
            return sep

        hdr_row.addWidget(_make_field("SHOP",        header.get("section", "—")), 2)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("TANGGAL",     header.get("date", "—")), 1)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("HARI",        _hari), 1)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("SHIFT",       header.get("shift", "—")), 1)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("HOUR",        _fmt(total_hour)), 1)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("OT",          overtime), 1)
        hdr_row.addWidget(_vsep_h())
        hdr_row.addWidget(_make_field("KOORDINATOR", header.get("coordinator", "—")), 2)
        outer_lyt.addWidget(hdr_bar)

        # ── SCROLL AREA ───────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main_lyt = QVBoxLayout(container)
        main_lyt.setContentsMargins(12, 12, 12, 12)
        main_lyt.setSpacing(8)

        # ── TOP ROW (3:3:2) ───────────────────────────────────────────────
        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        # — Production Volume panel —
        card_prod = _card()
        lay_prod = QVBoxLayout(card_prod)
        lay_prod.setContentsMargins(10, 10, 10, 10)
        lay_prod.setSpacing(6)
        lay_prod.addWidget(_hdr_lbl("Production Volume"))
        tbl_prod = _tbl(
            5,
            ["Model", "Plan Qty", "Act Qty", "Plan H", "Act H"],
            fixed_cols=[(0, 80), (1, 65), (2, 65), (3, 65), (4, 65)],
        )
        tbl_prod.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        for i, p in enumerate(produksi):
            tbl_prod.insertRow(i)
            tbl_prod.setRowHeight(i, 28)
            tbl_prod.setItem(i, 0, _c(p.get("model", "")))
            tbl_prod.setItem(i, 1, _c(f"{p.get('plan_unit', 0) or 0:.0f}", Qt.AlignCenter))
            tbl_prod.setItem(i, 2, _c(f"{p.get('actual_unit', 0) or 0:.0f}", Qt.AlignCenter))
            tbl_prod.setItem(i, 3, _c(_fmt(p.get("plan_whour", 0) or 0), Qt.AlignCenter))
            tbl_prod.setItem(i, 4, _c(_fmt(p.get("actual_whour", 0) or 0), Qt.AlignCenter))
        tbl_prod.setMinimumHeight(110)
        lay_prod.addWidget(tbl_prod)
        top_row.addWidget(card_prod, 3)

        # — Absence panel —
        card_abs = _card()
        lay_abs = QVBoxLayout(card_abs)
        lay_abs.setContentsMargins(10, 10, 10, 10)
        lay_abs.setSpacing(6)
        lay_abs.addWidget(_hdr_lbl("Absence"))
        tbl_abs = _tbl(
            4,
            ["NIK", "Name", "Note", "Hour"],
            fixed_cols=[(0, 70), (3, 50)],
        )
        tbl_abs.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        tbl_abs.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        for i, a in enumerate(absen):
            tbl_abs.insertRow(i)
            tbl_abs.setRowHeight(i, 28)
            tbl_abs.setItem(i, 0, _c(a.get("nik", ""), Qt.AlignCenter))
            tbl_abs.setItem(i, 1, _c(a.get("nama", "")))
            tbl_abs.setItem(i, 2, _c(a.get("shop", "")))
            tbl_abs.setItem(i, 3, _c(a.get("keterangan", ""), Qt.AlignCenter))
        tbl_abs.setMinimumHeight(110)
        lay_abs.addWidget(tbl_abs)
        top_row.addWidget(card_abs, 3)

        # — Calculation Hour panel —
        card_calc = _card()
        lay_calc = QVBoxLayout(card_calc)
        lay_calc.setContentsMargins(12, 10, 12, 10)
        lay_calc.setSpacing(0)
        lay_calc.addWidget(_hdr_lbl("Calculation Hour"))
        lay_calc.addSpacing(8)

        grid_calc = QGridLayout()
        grid_calc.setSpacing(3)
        grid_calc.setColumnMinimumWidth(0, 90)

        def _rlbl(t):
            l = QLabel(t); l.setStyleSheet(_FLD_LBL_D); return l

        def _vlbl(t):
            l = QLabel(t)
            l.setStyleSheet("color:#ffffff; font-size:11px; font-weight:bold;")
            l.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return l

        calc_rows = [
            ("Process",     _fmt(process)),
            ("Preparation", _fmt(prep_h)),
            ("Quality",     "0"),
            ("Line Stop",   _fmt(linestop)),
            ("Absence",     _fmt(absence)),
            ("Other",       _fmt(other_h)),
        ]
        for i, (name, val) in enumerate(calc_rows):
            grid_calc.addWidget(_rlbl(name), i, 0)
            grid_calc.addWidget(_vlbl(val),  i, 1)

        div = QFrame()
        div.setFrameShape(QFrame.HLine)
        div.setFixedHeight(1)
        div.setStyleSheet("background-color:#303030; border:none;")
        grid_calc.addWidget(div, len(calc_rows), 0, 1, 2)

        lbl_tot = QLabel("TOTAL")
        lbl_tot.setStyleSheet("color:#969696; font-size:11px; font-weight:bold;")
        grid_calc.addWidget(lbl_tot, len(calc_rows) + 1, 0)
        grid_calc.addWidget(_vlbl(_fmt(total_hour)), len(calc_rows) + 1, 1)

        lay_calc.addLayout(grid_calc)
        lay_calc.addSpacing(6)

        bal_frame = QFrame()
        bal_frame.setObjectName("balFrame")
        bal_frame.setStyleSheet(
            "#balFrame { background-color:#1e1e1e; border:1px solid #303030;"
            " border-left:3px solid #da291c; }"
        )
        bal_row = QHBoxLayout(bal_frame)
        bal_row.setContentsMargins(8, 5, 8, 5)
        lbl_bal_title = QLabel("BALANCE")
        lbl_bal_title.setStyleSheet("color:#969696; font-size:11px; font-weight:bold;")
        ok  = abs(balance) < 0.001
        clr = "rgb(80,200,100)" if ok else "rgb(220,80,80)"
        lbl_bal_val = QLabel(_fmt(balance))
        lbl_bal_val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lbl_bal_val.setStyleSheet(f"color:{clr}; font-size:13px; font-weight:bold;")
        bal_row.addWidget(lbl_bal_title)
        bal_row.addStretch()
        bal_row.addWidget(lbl_bal_val)
        lay_calc.addWidget(bal_frame)
        lay_calc.addStretch()
        top_row.addWidget(card_calc, 2)

        main_lyt.addLayout(top_row)

        # ── INHOUSE CLAIM ─────────────────────────────────────────────────
        card_ic = _card()
        lay_ic  = QVBoxLayout(card_ic)
        lay_ic.setContentsMargins(10, 10, 10, 10)
        lay_ic.setSpacing(6)
        lay_ic.addWidget(_hdr_lbl("In House (Reject) and Market Claim"))
        tbl_ic = _tbl(
            12,
            ["No", "Model", "OP/ST", "Item", "Qty", "Satuan",
             "Cause / Penyebab", "Action / Perbaikan", "Factor",
             "Stop (H)", "Lost (H)", "Status"],
            fixed_cols=[(0, 28), (1, 60), (2, 60), (4, 42),
                        (5, 55), (8, 75), (9, 50), (10, 50), (11, 65)],
        )
        tbl_ic.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        tbl_ic.horizontalHeader().setSectionResizeMode(6, QHeaderView.Stretch)
        tbl_ic.horizontalHeader().setSectionResizeMode(7, QHeaderView.Stretch)
        tbl_ic.setWordWrap(True)
        _wrap_ic = _WrapDelegate(tbl_ic)
        for _col in (3, 6, 7):
            tbl_ic.setItemDelegateForColumn(_col, _wrap_ic)
        tbl_ic.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        for i, ic in enumerate(inhouse_claim):
            tbl_ic.insertRow(i)
            tbl_ic.setItem(i, 0,  _c(str(i + 1), Qt.AlignCenter))
            tbl_ic.setItem(i, 1,  _c(ic.get("model", ""), Qt.AlignCenter))
            tbl_ic.setItem(i, 2,  _c(ic.get("op_no_st", ""), Qt.AlignCenter))
            tbl_ic.setItem(i, 3,  _c(ic.get("item", "")))
            tbl_ic.setItem(i, 4,  _c(f"{ic.get('qty', 0) or 0:.0f}", Qt.AlignCenter))
            tbl_ic.setItem(i, 5,  _c(ic.get("satuan", ""), Qt.AlignCenter))
            tbl_ic.setItem(i, 6,  _c(ic.get("penyebab", "")))
            tbl_ic.setItem(i, 7,  _c(ic.get("tindakan", "")))
            tbl_ic.setItem(i, 8,  _c(ic.get("faktor", ""), Qt.AlignCenter))
            tbl_ic.setItem(i, 9,  _c(_fmt(ic.get("stop_hr", 0) or 0), Qt.AlignCenter))
            tbl_ic.setItem(i, 10, _c(_fmt(ic.get("lost_hr", 0) or 0), Qt.AlignCenter))
            st = ic.get("status", "").upper()
            st_item = _c(ic.get("status", ""), Qt.AlignCenter)
            if st == "NG":
                st_item.setForeground(QColor(220, 100, 100))
            elif st == "PENDING":
                st_item.setForeground(QColor(220, 170, 80))
            elif st == "OK":
                st_item.setForeground(QColor(80, 200, 100))
            tbl_ic.setItem(i, 11, st_item)
        tbl_ic.setMinimumHeight(120)
        lay_ic.addWidget(tbl_ic)
        main_lyt.addWidget(card_ic)

        # ── LINE STOP ─────────────────────────────────────────────────────
        card_ls = _card()
        lay_ls  = QVBoxLayout(card_ls)
        lay_ls.setContentsMargins(10, 10, 10, 10)
        lay_ls.setSpacing(6)
        lay_ls.addWidget(_hdr_lbl(
            "Line Stop  ( Tool / Model Change / Man / Machine / Material / Meeting / Quality and Others )"
        ))
        tbl_ls = _tbl(
            11,
            ["No", "Model", "OP/ST", "Problem / Masalah",
             "Cause / Penyebab", "Action / Perbaikan",
             "Factor", "Start", "End", "Stop", "Lost"],
            fixed_cols=[(0, 28), (1, 60), (2, 60), (6, 80),
                        (7, 54), (8, 54), (9, 52), (10, 52)],
        )
        tbl_ls.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
        tbl_ls.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        tbl_ls.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
        tbl_ls.setWordWrap(True)
        _wrap_ls = _WrapDelegate(tbl_ls)
        for _col in (3, 4, 5):
            tbl_ls.setItemDelegateForColumn(_col, _wrap_ls)
        tbl_ls.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        import re as _re
        for i, c in enumerate(catatan):
            desc = c.get("description", "") or ""
            m    = _re.match(r"^\[([^\]]+)\]", desc)
            model_val = m.group(1) if m else ""
            tbl_ls.insertRow(i)
            tbl_ls.setItem(i, 0,  _c(str(i + 1), Qt.AlignCenter))
            tbl_ls.setItem(i, 1,  _c(model_val, Qt.AlignCenter))
            tbl_ls.setItem(i, 2,  _c(c.get("ra_number", ""), Qt.AlignCenter))
            tbl_ls.setItem(i, 3,  _c(desc))
            tbl_ls.setItem(i, 4,  _c(c.get("cause", "")))
            tbl_ls.setItem(i, 5,  _c(c.get("corrective_action", "")))
            tbl_ls.setItem(i, 6,  _c(c.get("category", ""), Qt.AlignCenter))
            tbl_ls.setItem(i, 7,  _c(c.get("start_time", ""), Qt.AlignCenter))
            tbl_ls.setItem(i, 8,  _c(c.get("end_time", ""), Qt.AlignCenter))
            tbl_ls.setItem(i, 9,  _c(_fmt(c.get("down_time", 0) or 0), Qt.AlignCenter))
            tbl_ls.setItem(i, 10, _c(_fmt(c.get("loss_time", 0) or 0), Qt.AlignCenter))
        tbl_ls.setMinimumHeight(120)
        lay_ls.addWidget(tbl_ls)
        main_lyt.addWidget(card_ls)

        # ── MATERIAL USED (kondisional) ───────────────────────────────────
        if materials:
            card_mat = _card()
            lay_mat  = QVBoxLayout(card_mat)
            lay_mat.setContentsMargins(10, 10, 10, 10)
            lay_mat.setSpacing(6)
            lay_mat.addWidget(_hdr_lbl("Material Used"))
            tbl_mat = _tbl(
                6,
                ["No", "Material Name", "Mat. No", "Qty", "Satuan", "Keterangan"],
                fixed_cols=[(0, 28), (2, 90), (3, 55), (4, 65)],
            )
            tbl_mat.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
            tbl_mat.horizontalHeader().setSectionResizeMode(5, QHeaderView.Stretch)
            for i, mat in enumerate(materials):
                tbl_mat.insertRow(i)
                tbl_mat.setRowHeight(i, 28)
                tbl_mat.setItem(i, 0, _c(str(i + 1), Qt.AlignCenter))
                tbl_mat.setItem(i, 1, _c(mat.get("material_name", "")))
                tbl_mat.setItem(i, 2, _c(mat.get("material_no", ""), Qt.AlignCenter))
                tbl_mat.setItem(i, 3, _c(f"{mat.get('qty', 0) or 0:.0f}", Qt.AlignCenter))
                tbl_mat.setItem(i, 4, _c(mat.get("satuan", ""), Qt.AlignCenter))
                tbl_mat.setItem(i, 5, _c(mat.get("keterangan", "")))
            tbl_mat.setMinimumHeight(min(len(materials) * 28 + 32, 200))
            lay_mat.addWidget(tbl_mat)
            main_lyt.addWidget(card_mat)

        main_lyt.addStretch()
        scroll.setWidget(container)
        outer_lyt.addWidget(scroll)

        # ── FOOTER ────────────────────────────────────────────────────────
        btn_close = QPushButton("Tutup")
        btn_close.setFixedHeight(34)
        btn_close.setStyleSheet(
            "QPushButton { background-color: #252525; color: #969696;"
            " border: none; border-radius: 0px; padding: 0 24px; font-size: 11px;"
            " border-top: 1px solid #303030; }"
            "QPushButton:hover { background-color: #303030; color: #ffffff; }"
        )
        btn_close.clicked.connect(dlg.close)
        footer = QHBoxLayout()
        footer.setContentsMargins(12, 8, 12, 8)
        footer.addStretch()
        footer.addWidget(btn_close)
        outer_lyt.addLayout(footer)

        dlg.exec()

    def _edit(self, report_id):
        QMessageBox.information(self, "Edit", f"Fitur edit laporan #{report_id} akan segera tersedia.")

    def _hapus(self, report_id):
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus laporan #{report_id}?\nSemua catatan masalah terkait juga akan dihapus.\nTindakan ini tidak dapat dibatalkan.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            ok, msg = hapus_laporan(report_id)
            if ok:
                QMessageBox.information(self, "Berhasil", msg)
                self.load_data_harian()
            else:
                QMessageBox.critical(self, "Gagal", msg)

    def _export_rah(self):
        if self.combo_section.currentData() is None:
            QMessageBox.warning(self, "Peringatan", "Pilih Shop terlebih dahulu sebelum export.")
            return
        if self.tabel_harian.rowCount() == 0:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor.")
            return

        all_catatan  = []
        all_manpower = []
        section_name = ""
        shift_name   = ""
        coordinator  = ""

        try:
            for row in getattr(self, "_loaded_rows", []):
                _, _, catatan, manpower, _, _, _ = get_detail_laporan(row["id"])
                all_catatan.extend(catatan)
                all_manpower.extend(manpower)
                if not section_name:
                    section_name = row.get("section", "")
                if not shift_name:
                    shift_name = row.get("shift", "")
                if not coordinator:
                    coordinator = row.get("coordinator", "")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat detail laporan: {e}")
            return

        header_info = {
            "date":        self.date_from.date().toString("yyyy-MM-dd"),
            "section":     section_name,
            "shift":       shift_name,
            "coordinator": coordinator,
            "approved_by": "",
            "checked_by":  "",
        }

        if not all_catatan:
            QMessageBox.warning(self, "Peringatan", "Tidak ada catatan masalah untuk diekspor.")
            return

        try:
            filepath = export_loss_time_record(
                header=header_info,
                catatan=all_catatan,
                produksi=[],
                manpower=all_manpower,
                absen=[],
                inhouse_claim=[],
            )
            QMessageBox.information(self, "Export Berhasil", f"File disimpan di:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Export", f"Gagal mengekspor: {e}")

    # Tab 2: Rekap Bulanan

    def _build_tab_rekap(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_section_rekap = QComboBox()
        self.combo_section_rekap.setMinimumHeight(30)
        self.combo_section_rekap.setMinimumWidth(155)
        self.combo_section_rekap.setStyleSheet(_COMBO_STYLE)
        self.combo_section_rekap.addItem("Semua", None)
        fl.addWidget(self.combo_section_rekap)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_bulan = QComboBox()
        self.combo_bulan.setMinimumHeight(30)
        self.combo_bulan.setMinimumWidth(115)
        self.combo_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate([
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember",
        ], 1):
            self.combo_bulan.addItem(name, i)
        self.combo_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_tahun = QLineEdit()
        self.input_tahun.setMinimumHeight(30)
        self.input_tahun.setFixedWidth(70)
        self.input_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_tahun)

        fl.addWidget(self._flabel("Factor"))
        self.combo_factor = QComboBox()
        self.combo_factor.setMinimumHeight(30)
        self.combo_factor.setMinimumWidth(140)
        self.combo_factor.setStyleSheet(_COMBO_STYLE)
        self.combo_factor.addItem("Semua")
        try:
            for name in get_all_category_names():
                self.combo_factor.addItem(name)
        except Exception:
            pass
        fl.addWidget(self.combo_factor)

        fl.addStretch()

        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_rekap_bulanan)
        fl.addWidget(btn_tampil)

        btn_export_r = QPushButton("Export Excel")
        btn_export_r.setMinimumSize(100, 32)
        btn_export_r.setStyleSheet(_BTN_EXPORT)
        btn_export_r.clicked.connect(self._export_rekap)
        fl.addWidget(btn_export_r)

        main.addWidget(card_f)

        # Table card
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        self.lbl_info_r = QLabel("Pilih filter dan klik Tampilkan.")
        self.lbl_info_r.setStyleSheet("color: #969696; font-size: 10px;")
        tl.addWidget(self.lbl_info_r)

        self.tabel_rekap = QTableWidget()
        self.tabel_rekap.setColumnCount(9)
        self.tabel_rekap.setHorizontalHeaderLabels([
            "No", "Date", "OPNo/St", "Problem/Masalah", "Cause/Penyebab",
            "Action/Perbaikan", "Factor", "Lost (H)", "Stop (H)",
        ])
        hh = self.tabel_rekap.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        for col in (3, 4, 5):
            hh.setSectionResizeMode(col, QHeaderView.Stretch)
        for col in (6, 7, 8):
            hh.setSectionResizeMode(col, QHeaderView.Fixed)
        self.tabel_rekap.verticalHeader().setVisible(False)
        self.tabel_rekap.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_rekap.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_rekap.setAlternatingRowColors(True)
        self.tabel_rekap.setStyleSheet(_TABLE_STYLE)
        self.tabel_rekap.setMinimumHeight(400)
        self.tabel_rekap.setColumnWidth(0, 40)
        self.tabel_rekap.setColumnWidth(1, 92)
        self.tabel_rekap.setColumnWidth(2, 80)
        self.tabel_rekap.setColumnWidth(6, 110)
        self.tabel_rekap.setColumnWidth(7, 80)
        self.tabel_rekap.setColumnWidth(8, 80)
        tl.addWidget(self.tabel_rekap)

        # Summary bar
        sum_frame = QFrame()
        sum_frame.setStyleSheet("QFrame { background-color: #2e2e2e; border-radius: 0px; }")
        sum_lyt = QHBoxLayout(sum_frame)
        sum_lyt.setContentsMargins(16, 10, 16, 10)
        sum_lyt.setSpacing(30)

        self.lbl_total_hour = QLabel("Total Working Hour: —")
        self.lbl_total_hour.setStyleSheet("color: #ffffff; font-size: 12px; font-weight: bold;")
        self.lbl_total_lost = QLabel("Total Lost: —")
        self.lbl_total_lost.setStyleSheet("color: #ffffff; font-size: 12px; font-weight: bold;")
        self.lbl_ratio = QLabel("Ratio Loss: —")
        self.lbl_ratio.setStyleSheet("color: #969696; font-size: 12px; font-weight: bold;")
        sum_lyt.addWidget(self.lbl_total_hour)
        sum_lyt.addWidget(self.lbl_total_lost)
        sum_lyt.addWidget(self.lbl_ratio)
        sum_lyt.addStretch()
        tl.addWidget(sum_frame)

        main.addWidget(card_t)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_rekap_bulanan(self):
        section_id = self.combo_section_rekap.currentData()
        bulan      = self.combo_bulan.currentData()
        factor     = self.combo_factor.currentText()
        try:
            tahun = int(self.input_tahun.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Peringatan", "Tahun tidak valid.")
            return

        try:
            rows, total_hour = _db_display_line_stop(section_id, bulan, tahun, factor)
        except Exception as e:
            self.lbl_info_r.setText(f"Gagal memuat data: {e}")
            self.tabel_rekap.setRowCount(0)
            return

        total_lost = sum(float(r[6]) for r in rows)
        ratio_loss = (total_lost / total_hour * 100) if total_hour > 0 else 0.0

        self.tabel_rekap.setRowCount(0)
        self.lbl_info_r.setText(
            f"{len(rows)} catatan — {self.combo_bulan.currentText()} {tahun}"
        )

        def _it(text, align=Qt.AlignLeft | Qt.AlignVCenter):
            it = QTableWidgetItem(str(text) if text is not None else "")
            it.setTextAlignment(align)
            return it

        for i, r in enumerate(rows):
            self.tabel_rekap.insertRow(i)
            self.tabel_rekap.setItem(i, 0, _it(i + 1, Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 1, _it(str(r[0]) if r[0] else "", Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 2, _it(r[1] or ""))
            self.tabel_rekap.setItem(i, 3, _it(r[2] or ""))
            self.tabel_rekap.setItem(i, 4, _it(r[3] or ""))
            self.tabel_rekap.setItem(i, 5, _it(r[4] or ""))
            self.tabel_rekap.setItem(i, 6, _it(r[5] or ""))
            self.tabel_rekap.setItem(i, 7, _it(f"{float(r[6]):.2f}", Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 8, _it(f"{float(r[7]):.2f}", Qt.AlignCenter))
            self.tabel_rekap.setRowHeight(i, 32)

        self.lbl_total_hour.setText(f"Total Working Hour: {total_hour:.2f} H")
        self.lbl_total_lost.setText(f"Total Lost: {total_lost:.2f} H")
        ratio_color = "#27ae60" if ratio_loss < 20 else "#f39c12" if ratio_loss < 40 else "#da291c"
        self.lbl_ratio.setText(f"Ratio Loss: {ratio_loss:.2f}%")
        self.lbl_ratio.setStyleSheet(
            f"color: {ratio_color}; font-size: 12px; font-weight: bold;"
        )

    def _export_rekap(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl tidak terinstall.\nJalankan: pip install openpyxl")
            return

        row_count = self.tabel_rekap.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Display Line Stop"

        thin = Side(style="thin", color="3C4147")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1C2028")

        col_headers = [
            "No", "Date", "OPNo/St", "Problem/Masalah",
            "Cause/Penyebab", "Action/Perbaikan", "Factor", "Lost (H)", "Stop (H)",
        ]
        center_cols = {0, 1, 6, 7, 8}
        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = Font(color="9696A0", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for ri in range(row_count):
            for ci in range(9):
                item = self.tabel_rekap.item(ri, ci)
                val = item.text() if item else ""
                cell = ws.cell(row=ri + 2, column=ci + 1, value=val)
                cell.alignment = Alignment(
                    horizontal="center" if ci in center_cols else "left",
                    vertical="center",
                )
                cell.border = border

        sum_row = row_count + 3
        for col_offset, lbl in enumerate([self.lbl_total_hour, self.lbl_total_lost, self.lbl_ratio]):
            ws.cell(row=sum_row, column=col_offset + 1, value=lbl.text()).font = Font(
                bold=True, color="C8C8C8"
            )

        for ci, w in enumerate([6, 12, 10, 30, 30, 30, 16, 10, 10], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20

        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        filepath = os.path.join(
            downloads, f"display_line_stop_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        )
        try:
            wb.save(filepath)
            QMessageBox.information(self, "Export Berhasil", f"File disimpan di:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Export", f"Gagal menyimpan file:\n{e}")

    # Tab 3: NG & Pending

    # Tab 3: Produktivitas Bulanan

    def _build_tab_produktivitas(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame(); card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f); fl.setContentsMargins(16, 12, 16, 12); fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_prod_section = QComboBox()
        self.combo_prod_section.setMinimumHeight(30)
        self.combo_prod_section.setMinimumWidth(155)
        self.combo_prod_section.setStyleSheet(_COMBO_STYLE)
        self.combo_prod_section.addItem("Semua", None)
        fl.addWidget(self.combo_prod_section)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_prod_bulan = QComboBox()
        self.combo_prod_bulan.setMinimumHeight(30)
        self.combo_prod_bulan.setMinimumWidth(115)
        self.combo_prod_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate(
            ["Januari","Februari","Maret","April","Mei","Juni",
             "Juli","Agustus","September","Oktober","November","Desember"], 1
        ):
            self.combo_prod_bulan.addItem(name, i)
        self.combo_prod_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_prod_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_prod_tahun = QLineEdit()
        self.input_prod_tahun.setMinimumHeight(30)
        self.input_prod_tahun.setFixedWidth(70)
        self.input_prod_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_prod_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_prod_tahun)

        fl.addStretch()
        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_produktivitas)
        fl.addWidget(btn_tampil)
        main.addWidget(card_f)

        # Content row
        content = QHBoxLayout(); content.setSpacing(12)

        # Left: breakdown table
        card_left = QFrame(); card_left.setStyleSheet(_CARD_STYLE)
        ll = QVBoxLayout(card_left); ll.setContentsMargins(16, 14, 16, 16); ll.setSpacing(8)
        lbl_bkdn = QLabel("Breakdown Jam Kerja")
        lbl_bkdn.setStyleSheet(
            "color: #ffffff; font-size: 12px; font-weight: bold;"
            " border-left: 3px solid #da291c; padding-left: 8px;"
        )
        ll.addWidget(lbl_bkdn)

        self.tabel_prod_bkdn = QTableWidget()
        self.tabel_prod_bkdn.setColumnCount(2)
        self.tabel_prod_bkdn.setHorizontalHeaderLabels(["Kategori", "Jam (H)"])
        self.tabel_prod_bkdn.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabel_prod_bkdn.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.tabel_prod_bkdn.setColumnWidth(1, 90)
        self.tabel_prod_bkdn.verticalHeader().setVisible(False)
        self.tabel_prod_bkdn.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_prod_bkdn.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_prod_bkdn.setMinimumHeight(300)
        self.tabel_prod_bkdn.setStyleSheet(_TABLE_STYLE)
        ll.addWidget(self.tabel_prod_bkdn)
        content.addWidget(card_left, 3)

        # Right: ratio + summary
        card_right = QFrame(); card_right.setStyleSheet(_CARD_STYLE)
        rl = QVBoxLayout(card_right); rl.setContentsMargins(16, 14, 16, 16); rl.setSpacing(12)

        lbl_rt = QLabel("Ratio Produktivitas")
        lbl_rt.setStyleSheet("color: #969696; font-size: 11px; font-weight: bold;")
        lbl_rt.setAlignment(Qt.AlignCenter)
        rl.addWidget(lbl_rt)

        ratio_frame = QFrame()
        ratio_frame.setObjectName("ratioFrame")
        ratio_frame.setStyleSheet(
            "#ratioFrame { background-color: #1e1e1e; border: 2px solid #303030; }"
        )
        rfl = QVBoxLayout(ratio_frame); rfl.setContentsMargins(16, 24, 16, 24)
        self._lbl_ratio_big = QLabel("—")
        self._lbl_ratio_big.setStyleSheet(
            "color: #27ae60; font-size: 40px; font-weight: bold;"
        )
        self._lbl_ratio_big.setAlignment(Qt.AlignCenter)
        rfl.addWidget(self._lbl_ratio_big)
        rl.addWidget(ratio_frame)

        sum_frame = QFrame()
        sum_frame.setStyleSheet(
            "QFrame { background-color: #1e1e1e; border: 1px solid #303030; }"
        )
        sg = QGridLayout(sum_frame); sg.setContentsMargins(12, 10, 12, 10); sg.setSpacing(6)
        sg.setColumnStretch(1, 1)

        def _srow(row, label, attr):
            l = QLabel(label); l.setStyleSheet("color: #969696; font-size: 11px;")
            v = QLabel("—")
            v.setStyleSheet("color: #ffffff; font-size: 11px; font-weight: bold;")
            v.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            sg.addWidget(l, row, 0); sg.addWidget(v, row, 1)
            setattr(self, attr, v)

        _srow(0, "Total Hour",   "_lbl_prod_total")
        _srow(1, "Loss Time",    "_lbl_prod_bal")
        _srow(2, "Laporan",      "_lbl_prod_count")
        rl.addWidget(sum_frame)
        rl.addStretch()
        content.addWidget(card_right, 2)

        main.addLayout(content)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_produktivitas(self):
        section_id = self.combo_prod_section.currentData()
        bulan = self.combo_prod_bulan.currentData()
        try:
            tahun = int(self.input_prod_tahun.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Peringatan", "Tahun tidak valid.")
            return

        try:
            data = get_monthly_productivity(section_id, bulan, tahun)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data: {e}")
            return

        tbl = self.tabel_prod_bkdn
        tbl.setRowCount(0)

        loss_hour  = sum(c["hours"] for c in data["categories"])
        total_hour = data["total_hour"]

        def _grp_row(text):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it = QTableWidgetItem(text)
            it.setFlags(Qt.ItemIsEnabled)
            it.setBackground(QColor("#111111"))
            it.setForeground(QColor("#da291c"))
            f = it.font(); f.setBold(True); it.setFont(f)
            tbl.setItem(r, 0, it)
            it2 = QTableWidgetItem("")
            it2.setFlags(Qt.ItemIsEnabled)
            it2.setBackground(QColor("#111111"))
            tbl.setItem(r, 1, it2)
            tbl.setRowHeight(r, 26)

        def _subgrp_row(text):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it = QTableWidgetItem(f"  {text}")
            it.setFlags(Qt.ItemIsEnabled)
            it.setForeground(QColor("#606060"))
            f = it.font(); f.setItalic(True); it.setFont(f)
            tbl.setItem(r, 0, it)
            it2 = QTableWidgetItem("")
            it2.setFlags(Qt.ItemIsEnabled)
            tbl.setItem(r, 1, it2)
            tbl.setRowHeight(r, 20)

        def _data_row(name, hours, fg=None):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it_n = QTableWidgetItem(f"    {name}")
            it_n.setFlags(Qt.ItemIsEnabled)
            it_h = QTableWidgetItem(_fmt(hours))
            it_h.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_h.setFlags(Qt.ItemIsEnabled)
            if fg:
                it_n.setForeground(QColor(fg))
                it_h.setForeground(QColor(fg))
            tbl.setItem(r, 0, it_n)
            tbl.setItem(r, 1, it_h)
            tbl.setRowHeight(r, 28)

        # Produktif 
        _grp_row("PRODUKTIF")
        _data_row("Process", data["process_hour"], "#80c880")

        # Non-Produktif (line stop by category/group)

        groups: dict[str, list] = {}
        for cat in data["categories"]:
            groups.setdefault(cat["group"], []).append(cat)

        if groups:
            _grp_row("NON-PRODUKTIF (Line Stop)")
            for grp_name, cats in groups.items():
                _subgrp_row(grp_name)
                for cat in cats:
                    _data_row(cat["name"], cat["hours"])

        # Waktu Tetap 
        _grp_row("WAKTU TETAP")
        _data_row("Preparation", data["prep_hour"])
        _data_row("Other",       data["other_hour"])

        # Ketidakhadiran
        _grp_row("KETIDAKHADIRAN")
        _data_row("Absence", data["absence_hour"])

        # Divider + TOTAL
        r = tbl.rowCount()
        tbl.insertRow(r)
        for c in range(2):
            it = QTableWidgetItem()
            it.setFlags(Qt.ItemIsEnabled)
            it.setBackground(QColor("#404040"))
            tbl.setItem(r, c, it)
        tbl.setRowHeight(r, 2)

        r = tbl.rowCount()
        tbl.insertRow(r)
        it_n = QTableWidgetItem("TOTAL")
        it_n.setFlags(Qt.ItemIsEnabled)
        fn = it_n.font(); fn.setBold(True); it_n.setFont(fn)
        it_n.setForeground(QColor("#ffffff"))
        it_h = QTableWidgetItem(_fmt(total_hour))
        it_h.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        it_h.setFlags(Qt.ItemIsEnabled)
        fh = it_h.font(); fh.setBold(True); it_h.setFont(fh)
        it_h.setForeground(QColor("#ffffff"))
        tbl.setItem(r, 0, it_n)
        tbl.setItem(r, 1, it_h)
        tbl.setRowHeight(r, 30)

        # Summary
        balance = loss_hour

        ratio = (data["process_hour"] / total_hour * 100) if total_hour > 0 else 0.0
        ratio_color = (
            "#27ae60" if ratio >= 80
            else "#f39c12" if ratio >= 60
            else "#da291c"
        )
        self._lbl_ratio_big.setText(f"{ratio:.1f}%")
        self._lbl_ratio_big.setStyleSheet(
            f"color: {ratio_color}; font-size: 40px; font-weight: bold;"
        )
        self._lbl_prod_total.setText(f"{total_hour:.2f} H")

        bal_color = "#27ae60" if balance < 0.01 else "#da291c"
        self._lbl_prod_bal.setText(_fmt(balance))
        self._lbl_prod_bal.setStyleSheet(
            f"color: {bal_color}; font-size: 11px; font-weight: bold;"
        )
        self._lbl_prod_count.setText(f"{data['report_count']} laporan")

    @staticmethod
    def _flabel(text):
        lbl = QLabel(text)
        lbl.setStyleSheet("color: #969696; font-size: 11px;")
        return lbl

    # Tab 5: Volume Produksi

    def _build_tab_volume(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll_v = QScrollArea()
        scroll_v.setWidgetResizable(True)
        scroll_v.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_vol_section = QComboBox()
        self.combo_vol_section.setMinimumHeight(30)
        self.combo_vol_section.setMinimumWidth(155)
        self.combo_vol_section.setStyleSheet(_COMBO_STYLE)
        fl.addWidget(self.combo_vol_section)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_vol_bulan = QComboBox()
        self.combo_vol_bulan.setMinimumHeight(30)
        self.combo_vol_bulan.setMinimumWidth(115)
        self.combo_vol_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate([
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember",
        ], 1):
            self.combo_vol_bulan.addItem(name, i)
        self.combo_vol_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_vol_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_vol_tahun = QLineEdit()
        self.input_vol_tahun.setMinimumHeight(30)
        self.input_vol_tahun.setFixedWidth(70)
        self.input_vol_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_vol_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_vol_tahun)

        fl.addStretch()
        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_volume_produksi)
        fl.addWidget(btn_tampil)
        main.addWidget(card_f)

        # Table card with horizontal scroll
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        self.lbl_info_vol = QLabel("Pilih Shop dan klik Tampilkan.")
        self.lbl_info_vol.setStyleSheet("color: #969696; font-size: 10px;")
        tl.addWidget(self.lbl_info_vol)

        scroll_h = QScrollArea()
        scroll_h.setWidgetResizable(True)
        scroll_h.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_h.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_h.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        self.tabel_volume = QTableWidget()
        self.tabel_volume.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_volume.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_volume.verticalHeader().setVisible(False)
        self.tabel_volume.setStyleSheet(_TABLE_STYLE)
        self.tabel_volume.horizontalHeader().setDefaultSectionSize(28)
        self.tabel_volume.horizontalHeader().setMinimumSectionSize(20)
        self.tabel_volume.setMinimumHeight(300)

        scroll_h.setWidget(self.tabel_volume)
        tl.addWidget(scroll_h)
        main.addWidget(card_t)
        main.addStretch()
        scroll_v.setWidget(container)
        outer.addWidget(scroll_v)
        return tab

    def load_volume_produksi(self):
        section_id = self.combo_vol_section.currentData()
        if section_id is None:
            QMessageBox.warning(self, "Peringatan", "Pilih Shop terlebih dahulu.")
            return
        bulan = self.combo_vol_bulan.currentData()
        try:
            tahun = int(self.input_vol_tahun.text().strip())
        except ValueError:
            return

        try:
            result = get_production_volume(section_id, bulan, tahun)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data: {e}")
            return

        tbl  = self.tabel_volume
        days = result["days"]

        tbl.setColumnCount(2 + days + 1)
        headers = ["Model", "Shift"] + [str(d) for d in range(1, days + 1)] + ["TOTAL"]
        tbl.setHorizontalHeaderLabels(headers)

        hh = tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.Fixed)
        hh.setSectionResizeMode(1, QHeaderView.Fixed)
        hh.setSectionResizeMode(2 + days, QHeaderView.Fixed)

        tbl.setColumnWidth(0, 100)
        tbl.setColumnWidth(1, 90)
        tbl.setColumnWidth(2 + days, 70)

        tbl.setRowCount(0)

        row_idx = 0
        for model in result["models"]:
            for shift in result["shifts"]:
                tbl.insertRow(row_idx)
                tbl.setRowHeight(row_idx, 24)

                is_day = (shift == result["shifts"][0])
                row_bg = QColor("#252525") if is_day else QColor("#1a1a1a")

                tbl.setItem(row_idx, 0, _vol_item(model, bold=is_day))
                tbl.setItem(row_idx, 1, _vol_item(shift, center=True))

                key      = (model, shift)
                day_data = result["data"].get(key, {})
                for d in range(1, days + 1):
                    val = day_data.get(d)
                    it  = QTableWidgetItem(str(val) if val else "")
                    it.setTextAlignment(Qt.AlignCenter)
                    it.setFlags(Qt.ItemIsEnabled)
                    if val:
                        it.setBackground(QColor("#222222") if is_day else QColor("#1e1e1e"))
                    else:
                        it.setBackground(row_bg)
                        it.setForeground(QColor("#2e2e2e"))
                    tbl.setItem(row_idx, d + 1, it)

                total    = result["totals"].get(key, 0)
                it_total = QTableWidgetItem(str(total) if total else "")
                it_total.setTextAlignment(Qt.AlignCenter)
                it_total.setFlags(Qt.ItemIsEnabled)
                it_total.setForeground(QColor("#da291c"))
                fn = it_total.font(); fn.setBold(True); it_total.setFont(fn)
                it_total.setBackground(row_bg)
                tbl.setItem(row_idx, 2 + days, it_total)

                # Day Shift rows: set model/shift cell background
                for c in (0, 1):
                    it = tbl.item(row_idx, c)
                    if it:
                        it.setBackground(row_bg)

                row_idx += 1

        bulan_nama = self.combo_vol_bulan.currentText()
        self.lbl_info_vol.setText(
            f"{row_idx} baris — {bulan_nama} {tahun} | {self.combo_vol_section.currentText()}"
        )


# Module-level helpers

def _vol_item(text: str, bold: bool = False, center: bool = False) -> QTableWidgetItem:
    it = QTableWidgetItem(text)
    it.setTextAlignment(Qt.AlignCenter if center else Qt.AlignLeft | Qt.AlignVCenter)
    it.setFlags(Qt.ItemIsEnabled)
    if bold:
        f = it.font(); f.setBold(True); it.setFont(f)
    return it