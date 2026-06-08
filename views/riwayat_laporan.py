import os

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QGridLayout,
    QLabel, QFrame, QComboBox, QPushButton, QDateEdit,
    QScrollArea, QTableWidget, QTableWidgetItem, QHeaderView,
    QMessageBox, QDialog, QDialogButtonBox, QAbstractItemView,
    QTabWidget, QLineEdit,
)
from PySide6.QtCore import Qt, QDate, QThread, Signal
from PySide6.QtGui import QColor

from controllers.master_controller import (
    get_all_sections, get_all_shifts, get_all_category_names,
)
from controllers.laporan_controller import get_detail_laporan, hapus_laporan
from controllers.riwayat_controller import (
    get_riwayat_laporan, get_monthly_productivity,
    get_production_volume, get_display_line_stop, get_ng_pending,
)
from modules.export_excel import export_loss_time_record, export_inhouse_ng_pending


# Shared styles — Light Theme

_DATE_STYLE = """
    QDateEdit {
        background-color: #FFFFFF;
        border: 1px solid #D1D5DB;
        border-radius: 0px; padding-left: 8px;
        color: #212121; font-size: 11px;
    }
    QDateEdit:focus { border: 1px solid #E60012; }
    QDateEdit::drop-down {
        border: none; width: 22px;
        subcontrol-origin: padding; subcontrol-position: top right;
    }
"""

_COMBO_STYLE = """
    QComboBox {
        background-color: #FFFFFF;
        border: 1px solid #D1D5DB;
        border-radius: 0px; padding-left: 8px;
        color: #212121; font-size: 11px;
    }
    QComboBox:focus { border: 1px solid #E60012; }
    QComboBox::drop-down { border: none; width: 24px; }
    QComboBox QAbstractItemView {
        background-color: #FFFFFF; color: #212121;
        selection-background-color: #F3F4F6;
        border: 1px solid #D1D5DB;
    }
"""

_TABLE_STYLE = """
    QTableWidget {
        background-color: #FFFFFF;
        border: 1px solid #D1D5DB; border-radius: 0px;
        gridline-color: #F3F4F6;
    }
    QTableWidget::item {
        color: #212121; padding: 4px 8px;
        background-color: #FFFFFF;
        border-bottom: 1px solid #F3F4F6;
    }
    QTableWidget::item:alternate { background-color: #F9FAFB; }
    QTableWidget::item:selected { background-color: #F3F4F6; color: #212121; }
    QHeaderView::section {
        background-color: #F8F9FA; color: #6B7280;
        border: none; border-bottom: 1px solid #D1D5DB;
        border-right: 1px solid #E5E7EB;
        padding: 5px 8px; font-weight: bold; font-size: 10px;
        text-transform: uppercase; letter-spacing: 1px;
    }
"""

_TAB_STYLE = """
    QTabWidget::pane { background-color: transparent; border: none; }
    QTabBar::tab {
        background-color: #F3F4F6; color: #9CA3AF;
        padding: 8px 22px; margin-right: 2px;
        border-radius: 0px; font-size: 10px; font-weight: bold;
        letter-spacing: 1px; text-transform: uppercase;
    }
    QTabBar::tab:selected {
        background-color: #FFFFFF; color: #212121;
        border-bottom: 2px solid #E60012;
    }
    QTabBar::tab:hover:!selected {
        background-color: #E5E7EB; color: #212121;
    }
"""

_INPUT_STYLE = """
    QLineEdit {
        background-color: #FFFFFF;
        border: 1px solid #D1D5DB; border-radius: 0px;
        padding-left: 8px; color: #212121; font-size: 11px;
    }
    QLineEdit:focus { border: 1px solid #E60012; }
"""

_CARD_STYLE = "QFrame { background-color: #FFFFFF; border-radius: 0px; border: 1px solid #E5E7EB; }"

_BTN_CARI = """
    QPushButton {
        background-color: #E60012; color: #ffffff;
        border: none; border-radius: 0px; font-size: 11px; padding: 0 12px;
        text-transform: uppercase; letter-spacing: 1px;
    }
    QPushButton:hover { background-color: #C0000F; }
"""

_BTN_RESET = """
    QPushButton {
        background-color: #F3F4F6; color: #6B7280;
        border: 1px solid #D1D5DB; border-radius: 0px; font-size: 11px; padding: 0 12px;
    }
    QPushButton:hover { background-color: #D1D5DB; color: #212121; }
"""

_BTN_EXPORT = """
    QPushButton {
        background-color: #E8F5E9; color: #2e7d32;
        border: 1px solid #A5D6A7; border-radius: 0px; font-size: 11px; padding: 0 12px;
    }
    QPushButton:hover { background-color: #C8E6C9; }
"""


#  DB helpers — thin wrappers ke controller (SQL ada di controllers/riwayat_controller.py)

def _db_display_line_stop(section_id, bulan, tahun, factor=None):
    result = get_display_line_stop(section_id, bulan, tahun, factor)
    rows = [
        (d["date"], d["ra_number"], d["description"], d["cause"],
         d["corrective_action"], d["factor"], d["loss_time"], d["down_time"])
        for d in result["details"]
    ]
    return rows, result["total_hour"]


def _db_ng_pending(section_id, date_from, date_to):
    result = get_ng_pending(section_id, date_from, date_to)
    ng = [
        (d["tanggal"], d["model"], d["op_no_st"], d["item"], d["qty"],
         d["penyebab"], d["tindakan"], d["faktor"], d["stop_hr"], d["lost_hr"], d["status"])
        for d in result["ng"]
    ]
    pending = [
        (d["tanggal"], d["model"], d["op_no_st"], d["item"], d["qty"],
         d["penyebab"], d["tindakan"], d["faktor"], d["stop_hr"], d["lost_hr"])
        for d in result["pending"]
    ]
    return ng, pending


# Background worker

class _RiwayatWorker(QThread):
    finished = Signal(list)
    error    = Signal(str)

    def __init__(self, section_id, shift_name, date_from, date_to):
        super().__init__()
        self.section_id = section_id
        self.shift_name = shift_name
        self.date_from  = date_from
        self.date_to    = date_to

    def run(self):
        try:
            rows = get_riwayat_laporan(
                section_id=self.section_id,
                shift_name=self.shift_name,
                date_from=self.date_from,
                date_to=self.date_to,
            )
            self.finished.emit(rows)
        except Exception as e:
            self.error.emit(str(e))


# Widget

class RiwayatLaporanWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._sections_loaded = False
        self._setup_ui()

    def showEvent(self, event):
        super().showEvent(event)
        self._load_sections_once()
        self.load_data_harian()

    def _load_sections_once(self):
        if self._sections_loaded:
            return
        try:
            sections = get_all_sections()
            # combo_section (Riwayat Harian) — tanpa item "Semua", shop wajib dipilih
            self.combo_section.blockSignals(True)
            for sid, sname in sections:
                self.combo_section.addItem(sname, sid)
            self.combo_section.blockSignals(False)
            # combo lain — dengan item "Semua"
            for combo in (self.combo_section_rekap, self.combo_section_ng, self.combo_prod_section):
                combo.blockSignals(True)
                for sid, sname in sections:
                    combo.addItem(sname, sid)
                combo.blockSignals(False)
            # combo volume produksi — tanpa "Semua", shop wajib dipilih
            self.combo_vol_section.blockSignals(True)
            for sid, sname in sections:
                self.combo_vol_section.addItem(sname, sid)
            self.combo_vol_section.blockSignals(False)
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", f"Gagal memuat daftar shop: {e}")
        try:
            shifts = get_all_shifts()
            self.combo_shift.blockSignals(True)
            self.combo_shift.clear()
            self.combo_shift.addItem("Semua", None)
            for s in shifts:
                self.combo_shift.addItem(s["name"])
            self.combo_shift.blockSignals(False)
        except Exception as e:
            QMessageBox.warning(self, "Peringatan", f"Gagal memuat daftar shift: {e}")
        self._sections_loaded = True

    def _on_tab_changed(self, idx):
        if idx == 1:
            self.load_rekap_bulanan()
        elif idx == 2:
            self.load_ng_pending()
        elif idx == 3:
            self.load_produktivitas()
        elif idx == 4:
            pass  # tidak auto-load, user harus pilih shop dulu

    def _setup_ui(self):
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet(_TAB_STYLE)
        self.tabs.currentChanged.connect(self._on_tab_changed)
        self.tabs.addTab(self._build_tab_harian(), "Riwayat Harian")
        self.tabs.addTab(self._build_tab_rekap(), "Rekap Bulanan")
        self.tabs.addTab(self._build_tab_ng_pending(), "NG & Pending")
        self.tabs.addTab(self._build_tab_produktivitas(), "Produktivitas")
        self.tabs.addTab(self._build_tab_volume(),       "Volume Produksi")
        outer.addWidget(self.tabs)

    # Tab 1: Riwayat Harian

    def _build_tab_harian(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_section = QComboBox()
        self.combo_section.setMinimumHeight(30)
        self.combo_section.setMinimumWidth(155)
        self.combo_section.setStyleSheet(_COMBO_STYLE)
        fl.addWidget(self.combo_section)

        fl.addWidget(self._flabel("Shift"))
        self.combo_shift = QComboBox()
        self.combo_shift.setMinimumHeight(30)
        self.combo_shift.setMinimumWidth(110)
        self.combo_shift.setStyleSheet(_COMBO_STYLE)
        self.combo_shift.addItem("Semua", None)
        fl.addWidget(self.combo_shift)

        fl.addWidget(self._flabel("Dari"))
        self.date_from = QDateEdit()
        self.date_from.setCalendarPopup(True)
        self.date_from.setMinimumHeight(30)
        self.date_from.setStyleSheet(_DATE_STYLE)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        fl.addWidget(self.date_from)

        fl.addWidget(self._flabel("Sampai"))
        self.date_to = QDateEdit()
        self.date_to.setCalendarPopup(True)
        self.date_to.setMinimumHeight(30)
        self.date_to.setStyleSheet(_DATE_STYLE)
        self.date_to.setDate(QDate.currentDate())
        fl.addWidget(self.date_to)

        fl.addStretch()

        btn_cari = QPushButton("Cari")
        btn_cari.setMinimumSize(76, 32)
        btn_cari.setStyleSheet(_BTN_CARI)
        btn_cari.clicked.connect(self.load_data_harian)
        fl.addWidget(btn_cari)

        btn_reset_f = QPushButton("Reset Filter")
        btn_reset_f.setMinimumSize(90, 32)
        btn_reset_f.setStyleSheet(_BTN_RESET)
        btn_reset_f.clicked.connect(self._reset_filter_harian)
        fl.addWidget(btn_reset_f)

        main.addWidget(card_f)

        # Table card
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        hdr_row = QHBoxLayout()
        self.lbl_info_h = QLabel("Memuat data...")
        self.lbl_info_h.setStyleSheet("color: #6B7280; font-size: 10px;")
        hdr_row.addWidget(self.lbl_info_h)
        hdr_row.addStretch()
        btn_export_rah = QPushButton("Export RAH")
        btn_export_rah.setMinimumSize(100, 30)
        btn_export_rah.setStyleSheet(_BTN_EXPORT)
        btn_export_rah.clicked.connect(self._export_rah)
        hdr_row.addWidget(btn_export_rah)
        tl.addLayout(hdr_row)

        self.tabel_harian = QTableWidget()
        self.tabel_harian.setColumnCount(7)
        self.tabel_harian.setHorizontalHeaderLabels([
            "No", "Tanggal", "Shop", "Shift", "Koordinator", "Jml Masalah", "Aksi",
        ])
        self.tabel_harian.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.tabel_harian.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self.tabel_harian.horizontalHeader().setSectionResizeMode(6, QHeaderView.Fixed)
        self.tabel_harian.horizontalHeader().setStretchLastSection(False)
        self.tabel_harian.horizontalHeader().setSortIndicatorShown(True)
        self.tabel_harian.verticalHeader().setVisible(False)
        self.tabel_harian.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_harian.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_harian.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tabel_harian.setSortingEnabled(True)
        self.tabel_harian.setAlternatingRowColors(True)
        self.tabel_harian.setStyleSheet(_TABLE_STYLE)
        self.tabel_harian.setColumnWidth(0, 40)
        self.tabel_harian.setColumnWidth(1, 92)
        self.tabel_harian.setColumnWidth(3, 92)
        self.tabel_harian.setColumnWidth(4, 135)
        self.tabel_harian.setColumnWidth(5, 90)
        self.tabel_harian.setColumnWidth(6, 310)
        self.tabel_harian.cellDoubleClicked.connect(self._on_row_double_clicked)
        tl.addWidget(self.tabel_harian)

        main.addWidget(card_t)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_data_harian(self):
        section_id = self.combo_section.currentData()
        shift_text = self.combo_shift.currentText()
        shift_name = None if shift_text == "Semua" else shift_text
        date_from  = self.date_from.date().toString("yyyy-MM-dd")
        date_to    = self.date_to.date().toString("yyyy-MM-dd")

        self.lbl_info_h.setText("Memuat data...")
        self.tabel_harian.setRowCount(0)

        self._worker = _RiwayatWorker(section_id, shift_name, date_from, date_to)
        self._worker.finished.connect(self._on_harian_loaded)
        self._worker.error.connect(lambda e: self.lbl_info_h.setText(f"Gagal: {e}"))
        self._worker.start()

    def _on_row_double_clicked(self, row, _col):
        """Double-click row → buka detail laporan langsung."""
        if row < len(self._loaded_rows):
            rid = self._loaded_rows[row]["id"]
            self._lihat(rid)

    def _on_harian_loaded(self, rows: list):
        self._loaded_rows = rows
        self.tabel_harian.setSortingEnabled(False)   # matikan selama insert
        self.tabel_harian.setRowCount(0)
        self.lbl_info_h.setText(f"{len(rows)} data ditemukan")

        def _item(text, align=Qt.AlignCenter):
            it = QTableWidgetItem(str(text) if text is not None else "")
            it.setTextAlignment(align)
            return it

        for i, row in enumerate(rows):
            self.tabel_harian.insertRow(i)
            self.tabel_harian.setItem(i, 0, _item(i + 1))
            self.tabel_harian.setItem(i, 1, _item(row["date"]))
            self.tabel_harian.setItem(i, 2, _item(row["section"], Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_harian.setItem(i, 3, _item(row["shift"]))
            self.tabel_harian.setItem(i, 4, _item(row["coordinator"], Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_harian.setItem(i, 5, _item(str(row["jml_masalah"])))

            report_id = row["id"]
            aksi_w = QWidget()
            aksi_w.setStyleSheet("background-color: #FFFFFF;")
            al = QHBoxLayout(aksi_w)
            al.setContentsMargins(8, 0, 8, 0)
            al.setSpacing(8)
            al.setAlignment(Qt.AlignVCenter | Qt.AlignCenter)

            btn_lihat = QPushButton("🔍 Lihat")
            btn_lihat.setMinimumWidth(70); btn_lihat.setFixedHeight(26)
            btn_lihat.setStyleSheet(
                "QPushButton { background-color: #E3F2FD; color: #1565C0;"
                " border: 1px solid #90CAF9; border-radius: 3px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #BBDEFB; }"
            )
            btn_lihat.clicked.connect(lambda _, rid=report_id: self._lihat(rid))

            btn_edit = QPushButton("✏ Edit")
            btn_edit.setMinimumWidth(58); btn_edit.setFixedHeight(26)
            btn_edit.setStyleSheet(
                "QPushButton { background-color: #FFF8E1; color: #F57F17;"
                " border: 1px solid #FFE082; border-radius: 3px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #FFE082; }"
            )
            btn_edit.clicked.connect(lambda _, rid=report_id: self._edit(rid))

            btn_hapus = QPushButton("🗑 Hapus")
            btn_hapus.setMinimumWidth(70); btn_hapus.setFixedHeight(26)
            btn_hapus.setStyleSheet(
                "QPushButton { background-color: #FFEBEE; color: #c62828;"
                " border: 1px solid #FFCDD2; border-radius: 3px; padding: 0 8px; font-size: 10px; }"
                "QPushButton:hover { background-color: #FFCDD2; }"
            )
            btn_hapus.clicked.connect(lambda _, rid=report_id: self._hapus(rid))

            al.addWidget(btn_lihat)
            al.addWidget(btn_edit)
            al.addWidget(btn_hapus)
            self.tabel_harian.setCellWidget(i, 6, aksi_w)
            self.tabel_harian.setRowHeight(i, 40)

        self.tabel_harian.setSortingEnabled(True)   # aktifkan kembali setelah insert

    def _reset_filter_harian(self):
        self.combo_section.setCurrentIndex(0)
        self.combo_shift.setCurrentIndex(0)
        self.date_from.setDate(QDate.currentDate().addMonths(-1))
        self.date_to.setDate(QDate.currentDate())
        self.load_data_harian()

    def _lihat(self, report_id):
        try:
            header, produksi, catatan, manpower, absen, inhouse_claim = get_detail_laporan(report_id)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat detail: {e}")
            return
        if header is None:
            QMessageBox.warning(self, "Tidak Ditemukan", f"Laporan #{report_id} tidak ditemukan.")
            return

        # ── Style constants — identik dengan input_laporan.py ───────────────
        _CARD = "QFrame { background-color: #FFFFFF; border-radius: 0px; border: 1px solid #E5E7EB; }"
        _TBL  = """
            QTableWidget { background-color: #FFFFFF; border: 1px solid #D1D5DB; gridline-color: #F3F4F6; }
            QTableWidget::item { color: #212121; padding: 3px 6px; background-color: #FFFFFF;
                border-bottom: 1px solid #F3F4F6; }
            QTableWidget::item:alternate { background-color: #F9FAFB; }
            QTableWidget::item:selected { background-color: #F3F4F6; color: #212121; }
            QHeaderView::section { background-color: #F8F9FA; color: #6B7280; border: none;
                border-bottom: 1px solid #D1D5DB; border-right: 1px solid #E5E7EB;
                padding: 4px 6px; font-weight: bold; font-size: 10px;
                text-transform: uppercase; letter-spacing: 1px; }
        """
        _HDR_LBL = ("color: #212121; font-size: 11px; font-weight: bold;"
                    " border-left: 2px solid #E60012; padding-left: 8px;"
                    " letter-spacing: 1px; text-transform: uppercase;")
        _FLD_LBL = "color: #6B7280; font-size: 10px; letter-spacing: 1px;"

        def _cell(val, align=Qt.AlignLeft | Qt.AlignVCenter):
            it = QTableWidgetItem(str(val) if val else "")
            it.setTextAlignment(align)
            return it

        def _setup_tbl(tbl):
            tbl.setStyleSheet(_TBL)
            tbl.verticalHeader().setVisible(False)
            tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectRows)

        # ── Dialog ──────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle(f"Detail Laporan — #{report_id}")
        dlg.setMinimumWidth(1060)
        dlg.setMinimumHeight(700)
        dlg.setStyleSheet("QDialog { background-color: #F8F9FA; color: #212121; }")

        outer = QVBoxLayout(dlg)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(0)

        # ── Header bar — identik dengan _build_header() di input_laporan ────
        hdr_bar = QFrame()
        hdr_bar.setStyleSheet(
            "QFrame { background-color: #FFFFFF; border-radius: 0px;"
            " border-top: 2px solid #E60012; border-bottom: 1px solid #E5E7EB; }"
        )
        hb = QHBoxLayout(hdr_bar)
        hb.setContentsMargins(12, 8, 12, 8)
        hb.setSpacing(8)

        def _vsep():
            s = QFrame(); s.setFrameShape(QFrame.VLine)
            s.setFixedWidth(1); s.setFixedHeight(22)
            s.setStyleSheet("background-color: #D1D5DB; border: none;")
            return s

        def _lbl(t):
            l = QLabel(t); l.setStyleSheet(_FLD_LBL); return l

        def _val(t):
            l = QLabel(str(t) if t else "—")
            l.setStyleSheet("color:#212121; font-size:11px; font-weight:bold;")
            return l

        hb.addWidget(_lbl("Shop"));       hb.addWidget(_val(header.get("section", "")))
        hb.addStretch(); hb.addWidget(_vsep()); hb.addStretch()
        hb.addWidget(_lbl("Date"));       hb.addWidget(_val(header.get("date", "")))
        hb.addStretch(); hb.addWidget(_vsep()); hb.addStretch()
        hb.addWidget(_lbl("Shift"));      hb.addWidget(_val(header.get("shift", "")))
        hb.addStretch(); hb.addWidget(_vsep()); hb.addStretch()

        plan_wh   = sum(p.get("plan_whour",   0) or 0 for p in produksi) if produksi else 0
        actual_wh = sum(p.get("actual_whour", 0) or 0 for p in produksi) if produksi else 0
        hb.addWidget(_lbl("Hour"));       hb.addWidget(_val(f"{actual_wh:.2f}"))
        hb.addStretch(); hb.addWidget(_vsep()); hb.addStretch()
        hb.addWidget(_lbl("OT"));         hb.addWidget(_val(header.get("overtime", "—")))
        hb.addStretch()
        lbl_user = QLabel(header.get("coordinator", "").upper())
        lbl_user.setStyleSheet(
            "color:#E60012; font-size:11px; font-weight:bold;"
            " padding: 3px 8px; background-color: #F8F9FA;"
        )
        hb.addWidget(lbl_user)
        outer.addWidget(hdr_bar)

        # ── Scrollable canvas ────────────────────────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(12, 12, 12, 12)
        main.setSpacing(8)

        # ════════════════════════════════════════════════════════════════════
        # TOP ROW  (3 panels) — Production | Absence | Manpower
        # identik dengan _build_top_row() di input_laporan
        # ════════════════════════════════════════════════════════════════════
        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        # ── Panel 1 : Production Volume ──────────────────────────────────────
        p1 = QFrame(); p1.setStyleSheet(_CARD)
        p1l = QVBoxLayout(p1)
        p1l.setContentsMargins(10, 10, 10, 10); p1l.setSpacing(6)
        p1l.addWidget(QLabel("Production Volume", styleSheet=_HDR_LBL))

        tbl_prod = QTableWidget(0, 5)
        tbl_prod.setHorizontalHeaderLabels(["Model", "Plan Qty", "Act Qty", "Plan H", "Act H"])
        h = tbl_prod.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.Stretch)
        tbl_prod.setMinimumHeight(130)
        _setup_tbl(tbl_prod)
        for i, p in enumerate(produksi):
            tbl_prod.insertRow(i)
            tbl_prod.setRowHeight(i, 30)
            tbl_prod.setItem(i, 0, _cell(p.get("model", "")))
            tbl_prod.setItem(i, 1, _cell(f"{p.get('plan_unit',0):.0f}",   Qt.AlignCenter))
            tbl_prod.setItem(i, 2, _cell(f"{p.get('actual_unit',0):.0f}", Qt.AlignCenter))
            tbl_prod.setItem(i, 3, _cell(f"{p.get('plan_whour',0):.4f}",  Qt.AlignCenter))
            tbl_prod.setItem(i, 4, _cell(f"{p.get('actual_whour',0):.4f}",Qt.AlignCenter))
        p1l.addWidget(tbl_prod)
        top_row.addWidget(p1, 3)

        # ── Panel 2 : Absence ────────────────────────────────────────────────
        p2 = QFrame(); p2.setStyleSheet(_CARD)
        p2l = QVBoxLayout(p2)
        p2l.setContentsMargins(10, 10, 10, 10); p2l.setSpacing(6)
        p2l.addWidget(QLabel("Absence", styleSheet=_HDR_LBL))

        tbl_absen = QTableWidget(0, 4)
        tbl_absen.setHorizontalHeaderLabels(["NIK", "Name", "Note", "Hour"])
        h = tbl_absen.horizontalHeader()
        h.setSectionResizeMode(0, QHeaderView.Fixed)
        h.setSectionResizeMode(1, QHeaderView.Stretch)
        h.setSectionResizeMode(2, QHeaderView.Fixed)
        h.setSectionResizeMode(3, QHeaderView.Fixed)
        tbl_absen.setColumnWidth(0, 55)
        tbl_absen.setColumnWidth(2, 80)
        tbl_absen.setColumnWidth(3, 52)
        tbl_absen.setMinimumHeight(130)
        _setup_tbl(tbl_absen)
        for i, a in enumerate(absen):
            tbl_absen.insertRow(i)
            tbl_absen.setRowHeight(i, 30)
            tbl_absen.setItem(i, 0, _cell(a.get("nik",        ""), Qt.AlignCenter))
            tbl_absen.setItem(i, 1, _cell(a.get("nama",       "")))
            tbl_absen.setItem(i, 2, _cell(a.get("keterangan", "")))
            tbl_absen.setItem(i, 3, _cell(""))   # hour tidak disimpan di absen
        p2l.addWidget(tbl_absen)
        top_row.addWidget(p2, 3)

        # ── Panel 3 : Calculation Hour (identik dengan _build_calc_hour_panel) ─
        prep_h   = header.get("preparation_min", 15.0) / 60
        sholat_h = header.get("sholat_min",      10.0) / 60
        process  = sum(p.get("actual_whour", 0) or 0 for p in produksi)
        linestop = sum(c.get("loss_time",    0) or 0 for c in catatan)
        absence  = 0.0   # jam absen tidak disimpan di tabel absen
        quality  = 0.0
        _ot_map  = {"-": 0.0, "2H": 2.0, "3H": 3.0, "11.4H": 11.4}
        ot_h     = _ot_map.get(header.get("overtime", "-"), 0.0)
        total    = header.get("shift_duration", 0.0) + ot_h
        balance  = total - process - prep_h - quality - linestop - absence - sholat_h

        p3 = QFrame(); p3.setStyleSheet(_CARD)
        p3l = QVBoxLayout(p3)
        p3l.setContentsMargins(12, 10, 12, 10); p3l.setSpacing(0)
        p3l.addWidget(QLabel("Calculation Hour", styleSheet=_HDR_LBL))
        p3l.addSpacing(8)

        def _rlbl(t):
            l = QLabel(t); l.setStyleSheet(_FLD_LBL); return l

        def _vlbl(t, muted=False):
            l = QLabel(str(t))
            clr = "#9CA3AF" if muted else "#212121"
            l.setStyleSheet(f"color:{clr}; font-size:11px; font-weight:bold; background:transparent;")
            l.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            return l

        grid_ch = QGridLayout()
        grid_ch.setSpacing(3)
        grid_ch.setColumnMinimumWidth(0, 80)

        rows_ch = [
            ("Process",     _vlbl(f"{process:.4f}")),
            ("Preparation", _vlbl(f"{prep_h:.4f}", muted=True)),
            ("Quality",     _vlbl(f"{quality:.4f}")),
            ("Line Stop",   _vlbl(f"{linestop:.4f}")),
            ("Absence",     _vlbl(f"{absence:.4f}")),
            ("Sholat",      _vlbl(f"{sholat_h:.4f}", muted=True)),
        ]
        for i, (name, wgt) in enumerate(rows_ch):
            grid_ch.addWidget(_rlbl(name), i, 0)
            grid_ch.addWidget(wgt, i, 1)

        div = QFrame(); div.setFrameShape(QFrame.HLine)
        div.setFixedHeight(1); div.setStyleSheet("background-color:#D1D5DB; border:none;")
        grid_ch.addWidget(div, len(rows_ch), 0, 1, 2)
        lbl_tot = QLabel("TOTAL")
        lbl_tot.setStyleSheet("color:#6B7280; font-size:11px; font-weight:bold; background:transparent;")
        lbl_tot_val = _vlbl(f"{total:.4f}")
        grid_ch.addWidget(lbl_tot,    len(rows_ch)+1, 0)
        grid_ch.addWidget(lbl_tot_val, len(rows_ch)+1, 1)

        p3l.addLayout(grid_ch)
        p3l.addSpacing(6)

        bal_frame = QFrame()
        bal_frame.setObjectName("balFrame")
        bal_frame.setStyleSheet(
            "#balFrame { background-color: #F8F9FA; border: 1px solid #D1D5DB;"
            " border-left: 3px solid #E60012; }"
        )
        bal_row = QHBoxLayout(bal_frame)
        bal_row.setContentsMargins(8, 5, 8, 5)
        lbl_bal = QLabel("BALANCE")
        lbl_bal.setStyleSheet("color:#6B7280; font-size:11px; font-weight:bold; background:transparent;")
        ok  = abs(balance) < 0.001
        clr = "rgb(80,200,100)" if ok else "rgb(220,80,80)"
        lbl_bal_val = QLabel(f"{balance:.4f}")
        lbl_bal_val.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        lbl_bal_val.setStyleSheet(f"color:{clr}; font-size:13px; font-weight:bold; background:transparent;")
        bal_row.addWidget(lbl_bal); bal_row.addStretch(); bal_row.addWidget(lbl_bal_val)
        p3l.addWidget(bal_frame)
        p3l.addStretch()
        top_row.addWidget(p3, 2)

        main.addLayout(top_row)

        # ════════════════════════════════════════════════════════════════════
        # INHOUSE CLAIM — full width, identik dengan _build_inhouse_claim()
        # ════════════════════════════════════════════════════════════════════
        ic_card = QFrame(); ic_card.setStyleSheet(_CARD)
        ic_lay  = QVBoxLayout(ic_card)
        ic_lay.setContentsMargins(10, 10, 10, 10); ic_lay.setSpacing(6)
        ic_lay.addWidget(QLabel("In House (Reject) and Market Claim", styleSheet=_HDR_LBL))

        tbl_ic = QTableWidget(0, 12)
        tbl_ic.setHorizontalHeaderLabels([
            "No", "Model", "OP/ST", "Item", "Qty", "Satuan",
            "Cause / Penyebab", "Action / Perbaikan", "Factor", "Hour", "Lost", "Status",
        ])
        h = tbl_ic.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.Fixed)
        h.setSectionResizeMode(3, QHeaderView.Stretch)
        h.setSectionResizeMode(6, QHeaderView.Stretch)
        h.setSectionResizeMode(7, QHeaderView.Stretch)
        tbl_ic.setColumnWidth(0,  28)
        tbl_ic.setColumnWidth(1,  60)
        tbl_ic.setColumnWidth(2,  60)
        tbl_ic.setColumnWidth(4,  42)
        tbl_ic.setColumnWidth(5,  55)
        tbl_ic.setColumnWidth(8,  75)
        tbl_ic.setColumnWidth(9,  50)
        tbl_ic.setColumnWidth(10, 50)
        tbl_ic.setColumnWidth(11, 75)
        tbl_ic.setWordWrap(True)
        tbl_ic.setMinimumHeight(140)
        _setup_tbl(tbl_ic)
        for i, ic in enumerate(inhouse_claim):
            tbl_ic.insertRow(i)
            tbl_ic.setRowHeight(i, 30)
            tbl_ic.setItem(i, 0,  _cell(str(i+1),                        Qt.AlignCenter))
            tbl_ic.setItem(i, 1,  _cell(ic.get("model",    "")))
            tbl_ic.setItem(i, 2,  _cell(ic.get("op_no_st", ""),          Qt.AlignCenter))
            tbl_ic.setItem(i, 3,  _cell(ic.get("item",     "")))
            tbl_ic.setItem(i, 4,  _cell(f"{ic.get('qty',0):.0f}",        Qt.AlignCenter))
            tbl_ic.setItem(i, 5,  _cell(ic.get("satuan",   ""),          Qt.AlignCenter))
            tbl_ic.setItem(i, 6,  _cell(ic.get("penyebab", "")))
            tbl_ic.setItem(i, 7,  _cell(ic.get("tindakan", "")))
            tbl_ic.setItem(i, 8,  _cell(ic.get("faktor",   ""),          Qt.AlignCenter))
            tbl_ic.setItem(i, 9,  _cell(f"{ic.get('stop_hr',0):.2f}",   Qt.AlignCenter))
            tbl_ic.setItem(i, 10, _cell(f"{ic.get('lost_hr',0):.2f}",   Qt.AlignCenter))
            s_val = ic.get("status", "")
            st_it = _cell(s_val, Qt.AlignCenter)
            su = s_val.upper()
            if su == "NG":        st_it.setForeground(QColor("#E60012"))
            elif su == "PENDING": st_it.setForeground(QColor("#FFC107"))
            elif su == "OK":      st_it.setForeground(QColor("#28A745"))
            tbl_ic.setItem(i, 11, st_it)
            tbl_ic.resizeRowToContents(i)
        ic_lay.addWidget(tbl_ic)
        main.addWidget(ic_card)

        # ════════════════════════════════════════════════════════════════════
        # LINE STOP — full width, identik dengan _build_line_stop()
        # ════════════════════════════════════════════════════════════════════
        ls_card = QFrame(); ls_card.setStyleSheet(_CARD)
        ls_lay  = QVBoxLayout(ls_card)
        ls_lay.setContentsMargins(10, 10, 10, 10); ls_lay.setSpacing(6)

        ls_title = QLabel(
            "Line Stop  ( Tool / Model Change / Man / Machine / Material / Meeting / Quality and Others )",
            styleSheet=_HDR_LBL,
        )
        ls_title.setWordWrap(True)
        ls_lay.addWidget(ls_title)

        tbl_ls = QTableWidget(0, 11)
        tbl_ls.setHorizontalHeaderLabels([
            "No", "No. RA", "Factor",
            "Problem / Masalah", "Cause / Penyebab", "Action / Perbaikan",
            "PIC", "Start", "End", "Stop", "Lost",
        ])
        h = tbl_ls.horizontalHeader()
        h.setSectionResizeMode(QHeaderView.Fixed)
        h.setSectionResizeMode(3, QHeaderView.Stretch)
        h.setSectionResizeMode(4, QHeaderView.Stretch)
        h.setSectionResizeMode(5, QHeaderView.Stretch)
        tbl_ls.setColumnWidth(0,  28)
        tbl_ls.setColumnWidth(1,  60)
        tbl_ls.setColumnWidth(2,  80)
        tbl_ls.setColumnWidth(6,  54)
        tbl_ls.setColumnWidth(7,  54)
        tbl_ls.setColumnWidth(8,  54)
        tbl_ls.setColumnWidth(9,  52)
        tbl_ls.setColumnWidth(10, 52)
        tbl_ls.setWordWrap(True)
        tbl_ls.setMinimumHeight(140)
        _setup_tbl(tbl_ls)
        for i, c in enumerate(catatan):
            tbl_ls.insertRow(i)
            tbl_ls.setItem(i, 0,  _cell(str(i+1),                          Qt.AlignCenter))
            tbl_ls.setItem(i, 1,  _cell(c.get("ra_number",        ""),     Qt.AlignCenter))
            tbl_ls.setItem(i, 2,  _cell(c.get("category",         "")))
            tbl_ls.setItem(i, 3,  _cell(c.get("description",      "")))
            tbl_ls.setItem(i, 4,  _cell(c.get("cause",            "")))
            tbl_ls.setItem(i, 5,  _cell(c.get("corrective_action","")))
            tbl_ls.setItem(i, 6,  _cell(c.get("pic",              ""),     Qt.AlignCenter))
            tbl_ls.setItem(i, 7,  _cell(c.get("start_time",       ""),     Qt.AlignCenter))
            tbl_ls.setItem(i, 8,  _cell(c.get("end_time",         ""),     Qt.AlignCenter))
            tbl_ls.setItem(i, 9,  _cell(f"{c.get('down_time',0):.2f}",    Qt.AlignCenter))
            tbl_ls.setItem(i, 10, _cell(f"{c.get('loss_time',0):.2f}",    Qt.AlignCenter))
            tbl_ls.resizeRowToContents(i)
        ls_lay.addWidget(tbl_ls)
        main.addWidget(ls_card)

        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll, 1)

        # ── Footer — identik dengan _build_footer() di input_laporan ─────────
        footer = QFrame()
        footer.setStyleSheet(
            "QFrame { background-color: #FFFFFF; border-top: 1px solid #E5E7EB; }"
        )
        f_lay = QHBoxLayout(footer)
        f_lay.setContentsMargins(12, 8, 12, 8)
        f_lay.setSpacing(10)
        f_lay.addStretch()
        btn_close = QPushButton("Tutup")
        btn_close.setMinimumSize(90, 34)
        btn_close.setStyleSheet(
            "QPushButton { background:#F3F4F6; color:#6B7280; border:none; font-size:11px; }"
            "QPushButton:hover { background:#D1D5DB; color:#212121; }"
        )
        btn_close.clicked.connect(dlg.close)
        f_lay.addWidget(btn_close)
        outer.addWidget(footer)

        dlg.exec()

    def _edit(self, report_id):
        QMessageBox.information(self, "Edit", f"Fitur edit laporan #{report_id} akan segera tersedia.")

    def _hapus(self, report_id):
        reply = QMessageBox.question(
            self, "Konfirmasi Hapus",
            f"Hapus laporan #{report_id}?\nSemua catatan masalah terkait juga akan dihapus.\nTindakan ini tidak dapat dibatalkan.",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )
        if reply == QMessageBox.Yes:
            ok, msg = hapus_laporan(report_id)
            if ok:
                QMessageBox.information(self, "Berhasil", msg)
                self.load_data_harian()
            else:
                QMessageBox.critical(self, "Gagal", msg)

    def _export_rah(self):
        if self.combo_section.currentData() is None:
            QMessageBox.warning(self, "Peringatan", "Pilih Shop terlebih dahulu sebelum export.")
            return
        if self.tabel_harian.rowCount() == 0:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor.")
            return

        all_catatan  = []
        all_manpower = []
        section_name = ""
        shift_name   = ""
        coordinator  = ""

        try:
            for row in getattr(self, "_loaded_rows", []):
                _, _, catatan, manpower, _, _ = get_detail_laporan(row["id"])
                all_catatan.extend(catatan)
                all_manpower.extend(manpower)
                if not section_name:
                    section_name = row.get("section", "")
                if not shift_name:
                    shift_name = row.get("shift", "")
                if not coordinator:
                    coordinator = row.get("coordinator", "")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat detail laporan: {e}")
            return

        header_info = {
            "date":        self.date_from.date().toString("yyyy-MM-dd"),
            "section":     section_name,
            "shift":       shift_name,
            "coordinator": coordinator,
            "approved_by": "",
            "checked_by":  "",
        }

        if not all_catatan:
            QMessageBox.warning(self, "Peringatan", "Tidak ada catatan masalah untuk diekspor.")
            return

        try:
            filepath = export_loss_time_record(
                header=header_info,
                catatan=all_catatan,
                produksi=[],
                manpower=all_manpower,
                absen=[],
                inhouse_claim=[],
            )
            QMessageBox.information(self, "Export Berhasil", f"File disimpan di:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Export", f"Gagal mengekspor: {e}")

    # Tab 2: Rekap Bulanan

    def _build_tab_rekap(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_section_rekap = QComboBox()
        self.combo_section_rekap.setMinimumHeight(30)
        self.combo_section_rekap.setMinimumWidth(155)
        self.combo_section_rekap.setStyleSheet(_COMBO_STYLE)
        self.combo_section_rekap.addItem("Semua", None)
        fl.addWidget(self.combo_section_rekap)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_bulan = QComboBox()
        self.combo_bulan.setMinimumHeight(30)
        self.combo_bulan.setMinimumWidth(115)
        self.combo_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate([
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember",
        ], 1):
            self.combo_bulan.addItem(name, i)
        self.combo_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_tahun = QLineEdit()
        self.input_tahun.setMinimumHeight(30)
        self.input_tahun.setFixedWidth(70)
        self.input_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_tahun)

        fl.addWidget(self._flabel("Factor"))
        self.combo_factor = QComboBox()
        self.combo_factor.setMinimumHeight(30)
        self.combo_factor.setMinimumWidth(140)
        self.combo_factor.setStyleSheet(_COMBO_STYLE)
        self.combo_factor.addItem("Semua")
        try:
            for name in get_all_category_names():
                self.combo_factor.addItem(name)
        except Exception:
            pass
        fl.addWidget(self.combo_factor)

        fl.addStretch()

        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_rekap_bulanan)
        fl.addWidget(btn_tampil)

        btn_export_r = QPushButton("Export Excel")
        btn_export_r.setMinimumSize(100, 32)
        btn_export_r.setStyleSheet(_BTN_EXPORT)
        btn_export_r.clicked.connect(self._export_rekap)
        fl.addWidget(btn_export_r)

        main.addWidget(card_f)

        # Table card
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        self.lbl_info_r = QLabel("Pilih filter dan klik Tampilkan.")
        self.lbl_info_r.setStyleSheet("color: #6B7280; font-size: 10px;")
        tl.addWidget(self.lbl_info_r)

        self.tabel_rekap = QTableWidget()
        self.tabel_rekap.setColumnCount(9)
        self.tabel_rekap.setHorizontalHeaderLabels([
            "No", "Date", "OPNo/St", "Problem/Masalah", "Cause/Penyebab",
            "Action/Perbaikan", "Factor", "Lost (H)", "Stop (H)",
        ])
        hh = self.tabel_rekap.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Interactive)
        for col in (3, 4, 5):
            hh.setSectionResizeMode(col, QHeaderView.Stretch)
        for col in (6, 7, 8):
            hh.setSectionResizeMode(col, QHeaderView.Fixed)
        self.tabel_rekap.verticalHeader().setVisible(False)
        self.tabel_rekap.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_rekap.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_rekap.setStyleSheet(_TABLE_STYLE)
        self.tabel_rekap.setMinimumHeight(400)
        self.tabel_rekap.setColumnWidth(0, 40)
        self.tabel_rekap.setColumnWidth(1, 92)
        self.tabel_rekap.setColumnWidth(2, 80)
        self.tabel_rekap.setColumnWidth(6, 110)
        self.tabel_rekap.setColumnWidth(7, 80)
        self.tabel_rekap.setColumnWidth(8, 80)
        tl.addWidget(self.tabel_rekap)

        # Summary bar
        sum_frame = QFrame()
        sum_frame.setStyleSheet("QFrame { background-color: #F8F9FA; border-radius: 0px; border: 1px solid #E5E7EB; }")
        sum_lyt = QHBoxLayout(sum_frame)
        sum_lyt.setContentsMargins(16, 10, 16, 10)
        sum_lyt.setSpacing(30)

        self.lbl_total_hour = QLabel("Total Working Hour: —")
        self.lbl_total_hour.setStyleSheet("color: #212121; font-size: 12px; font-weight: bold;")
        self.lbl_total_lost = QLabel("Total Lost: —")
        self.lbl_total_lost.setStyleSheet("color: #212121; font-size: 12px; font-weight: bold;")
        self.lbl_ratio = QLabel("Ratio Loss: —")
        self.lbl_ratio.setStyleSheet("color: #6B7280; font-size: 12px; font-weight: bold;")
        sum_lyt.addWidget(self.lbl_total_hour)
        sum_lyt.addWidget(self.lbl_total_lost)
        sum_lyt.addWidget(self.lbl_ratio)
        sum_lyt.addStretch()
        tl.addWidget(sum_frame)

        main.addWidget(card_t)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_rekap_bulanan(self):
        section_id = self.combo_section_rekap.currentData()
        bulan      = self.combo_bulan.currentData()
        factor     = self.combo_factor.currentText()
        try:
            tahun = int(self.input_tahun.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Peringatan", "Tahun tidak valid.")
            return

        try:
            rows, total_hour = _db_display_line_stop(section_id, bulan, tahun, factor)
        except Exception as e:
            self.lbl_info_r.setText(f"Gagal memuat data: {e}")
            self.tabel_rekap.setRowCount(0)
            return

        total_lost = sum(float(r[6]) for r in rows)
        ratio_loss = (total_lost / total_hour * 100) if total_hour > 0 else 0.0

        self.tabel_rekap.setRowCount(0)
        self.lbl_info_r.setText(
            f"{len(rows)} catatan — {self.combo_bulan.currentText()} {tahun}"
        )

        def _it(text, align=Qt.AlignLeft | Qt.AlignVCenter):
            it = QTableWidgetItem(str(text) if text is not None else "")
            it.setTextAlignment(align)
            return it

        for i, r in enumerate(rows):
            self.tabel_rekap.insertRow(i)
            self.tabel_rekap.setItem(i, 0, _it(i + 1, Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 1, _it(str(r[0]) if r[0] else "", Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 2, _it(r[1] or ""))
            self.tabel_rekap.setItem(i, 3, _it(r[2] or ""))
            self.tabel_rekap.setItem(i, 4, _it(r[3] or ""))
            self.tabel_rekap.setItem(i, 5, _it(r[4] or ""))
            self.tabel_rekap.setItem(i, 6, _it(r[5] or ""))
            self.tabel_rekap.setItem(i, 7, _it(f"{float(r[6]):.2f}", Qt.AlignCenter))
            self.tabel_rekap.setItem(i, 8, _it(f"{float(r[7]):.2f}", Qt.AlignCenter))
            self.tabel_rekap.setRowHeight(i, 32)

        self.lbl_total_hour.setText(f"Total Working Hour: {total_hour:.2f} H")
        self.lbl_total_lost.setText(f"Total Lost: {total_lost:.2f} H")
        ratio_color = "#27ae60" if ratio_loss < 20 else "#f39c12" if ratio_loss < 40 else "#E60012"
        self.lbl_ratio.setText(f"Ratio Loss: {ratio_loss:.2f}%")
        self.lbl_ratio.setStyleSheet(
            f"color: {ratio_color}; font-size: 12px; font-weight: bold;"
        )

    def _export_rekap(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            QMessageBox.critical(self, "Error", "openpyxl tidak terinstall.\nJalankan: pip install openpyxl")
            return

        row_count = self.tabel_rekap.rowCount()
        if row_count == 0:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Display Line Stop"

        thin = Side(style="thin", color="3C4147")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        hdr_fill = PatternFill("solid", fgColor="1C2028")

        col_headers = [
            "No", "Date", "OPNo/St", "Problem/Masalah",
            "Cause/Penyebab", "Action/Perbaikan", "Factor", "Lost (H)", "Stop (H)",
        ]
        center_cols = {0, 1, 6, 7, 8}
        for ci, h in enumerate(col_headers, 1):
            cell = ws.cell(row=1, column=ci, value=h)
            cell.fill = hdr_fill
            cell.font = Font(color="9696A0", bold=True, size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        for ri in range(row_count):
            for ci in range(9):
                item = self.tabel_rekap.item(ri, ci)
                val = item.text() if item else ""
                cell = ws.cell(row=ri + 2, column=ci + 1, value=val)
                cell.alignment = Alignment(
                    horizontal="center" if ci in center_cols else "left",
                    vertical="center",
                )
                cell.border = border

        sum_row = row_count + 3
        for col_offset, lbl in enumerate([self.lbl_total_hour, self.lbl_total_lost, self.lbl_ratio]):
            ws.cell(row=sum_row, column=col_offset + 1, value=lbl.text()).font = Font(
                bold=True, color="C8C8C8"
            )

        for ci, w in enumerate([6, 12, 10, 30, 30, 30, 16, 10, 10], 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[1].height = 20

        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        filepath = os.path.join(
            downloads, f"display_line_stop_{QDate.currentDate().toString('yyyyMMdd')}.xlsx"
        )
        try:
            wb.save(filepath)
            QMessageBox.information(self, "Export Berhasil", f"File disimpan di:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Export", f"Gagal menyimpan file:\n{e}")

    # Tab 3: NG & Pending

    def _build_tab_ng_pending(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_section_ng = QComboBox()
        self.combo_section_ng.setMinimumHeight(30)
        self.combo_section_ng.setMinimumWidth(155)
        self.combo_section_ng.setStyleSheet(_COMBO_STYLE)
        self.combo_section_ng.addItem("Semua", None)
        fl.addWidget(self.combo_section_ng)

        fl.addWidget(self._flabel("Dari"))
        self.date_ng_from = QDateEdit()
        self.date_ng_from.setCalendarPopup(True)
        self.date_ng_from.setMinimumHeight(30)
        self.date_ng_from.setStyleSheet(_DATE_STYLE)
        self.date_ng_from.setDate(QDate.currentDate().addMonths(-1))
        fl.addWidget(self.date_ng_from)

        fl.addWidget(self._flabel("Sampai"))
        self.date_ng_to = QDateEdit()
        self.date_ng_to.setCalendarPopup(True)
        self.date_ng_to.setMinimumHeight(30)
        self.date_ng_to.setStyleSheet(_DATE_STYLE)
        self.date_ng_to.setDate(QDate.currentDate())
        fl.addWidget(self.date_ng_to)

        fl.addStretch()

        btn_cari_ng = QPushButton("Cari")
        btn_cari_ng.setMinimumSize(76, 32)
        btn_cari_ng.setStyleSheet(_BTN_CARI)
        btn_cari_ng.clicked.connect(self.load_ng_pending)
        fl.addWidget(btn_cari_ng)

        btn_export_ng = QPushButton("Export Excel")
        btn_export_ng.setMinimumSize(100, 32)
        btn_export_ng.setStyleSheet(_BTN_EXPORT)
        btn_export_ng.clicked.connect(self._export_ng_pending)
        fl.addWidget(btn_export_ng)

        main.addWidget(card_f)

        # Inhouse Claim card
        card_ic = QFrame()
        card_ic.setStyleSheet(_CARD_STYLE)
        ic_lyt = QVBoxLayout(card_ic)
        ic_lyt.setContentsMargins(16, 16, 16, 16)
        ic_lyt.setSpacing(8)

        ic_hdr = QHBoxLayout()
        lbl_ic = QLabel("Inhouse Claim (NG)")
        lbl_ic.setStyleSheet(
            "color: #212121; font-size: 12px; font-weight: bold;"
            " border-left: 3px solid #E60012; padding-left: 8px;"
        )
        self.lbl_info_ic = QLabel("")
        self.lbl_info_ic.setStyleSheet("color: #6B7280; font-size: 10px;")
        ic_hdr.addWidget(lbl_ic)
        ic_hdr.addWidget(self.lbl_info_ic)
        ic_hdr.addStretch()
        ic_lyt.addLayout(ic_hdr)

        self.tabel_ic = QTableWidget()
        self.tabel_ic.setColumnCount(12)
        self.tabel_ic.setHorizontalHeaderLabels([
            "No", "Tanggal", "Model", "OP/St", "Item", "Qty",
            "Penyebab", "Tindakan", "Faktor", "Stop (H)", "Lost (H)", "Status",
        ])
        self.tabel_ic.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for fc in (0, 1, 2, 3, 5, 9, 10, 11):
            self.tabel_ic.horizontalHeader().setSectionResizeMode(
                fc, QHeaderView.ResizeToContents
            )
        self.tabel_ic.verticalHeader().setVisible(False)
        self.tabel_ic.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_ic.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_ic.setMinimumHeight(200)
        self.tabel_ic.setStyleSheet(_TABLE_STYLE)
        ic_lyt.addWidget(self.tabel_ic)

        main.addWidget(card_ic)

        # Part Pending card
        card_pp = QFrame()
        card_pp.setStyleSheet(_CARD_STYLE)
        pp_lyt = QVBoxLayout(card_pp)
        pp_lyt.setContentsMargins(16, 16, 16, 16)
        pp_lyt.setSpacing(8)

        pp_hdr = QHBoxLayout()
        lbl_pp = QLabel("Part Pending")
        lbl_pp.setStyleSheet(
            "color: #212121; font-size: 12px; font-weight: bold;"
            " border-left: 3px solid #E60012; padding-left: 8px;"
        )
        self.lbl_info_pp = QLabel("")
        self.lbl_info_pp.setStyleSheet("color: #6B7280; font-size: 10px;")
        pp_hdr.addWidget(lbl_pp)
        pp_hdr.addWidget(self.lbl_info_pp)
        pp_hdr.addStretch()
        pp_lyt.addLayout(pp_hdr)

        self.tabel_pp = QTableWidget()
        self.tabel_pp.setColumnCount(11)
        self.tabel_pp.setHorizontalHeaderLabels([
            "No", "Tanggal", "Model", "OP/St", "Item", "Qty",
            "Penyebab", "Tindakan", "Faktor", "Stop (H)", "Lost (H)",
        ])
        self.tabel_pp.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        for fc in (0, 1, 2, 3, 5, 9, 10):
            self.tabel_pp.horizontalHeader().setSectionResizeMode(
                fc, QHeaderView.ResizeToContents
            )
        self.tabel_pp.verticalHeader().setVisible(False)
        self.tabel_pp.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_pp.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_pp.setMinimumHeight(180)
        self.tabel_pp.setStyleSheet(_TABLE_STYLE)
        pp_lyt.addWidget(self.tabel_pp)

        main.addWidget(card_pp)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_ng_pending(self):
        section_id = self.combo_section_ng.currentData()
        date_from = self.date_ng_from.date().toString("yyyy-MM-dd")
        date_to = self.date_ng_to.date().toString("yyyy-MM-dd")

        try:
            inhouse_ng, inhouse_pending = _db_ng_pending(section_id, date_from, date_to)
        except Exception as e:
            self.lbl_info_ic.setText(f"Gagal memuat data: {e}")
            self.tabel_ic.setRowCount(0)
            self.tabel_pp.setRowCount(0)
            return

        def _item(text, align=Qt.AlignCenter):
            it = QTableWidgetItem(str(text) if text is not None else "")
            it.setTextAlignment(align)
            return it

        self.tabel_ic.setRowCount(0)
        self.lbl_info_ic.setText(f"— {len(inhouse_ng)} data")
        for i, row in enumerate(inhouse_ng):
            tanggal, model, op_no_st, item, qty, penyebab, tindakan, faktor, stop_hr, lost_hr, status = row
            self.tabel_ic.insertRow(i)
            self.tabel_ic.setItem(i, 0,  _item(i + 1))
            self.tabel_ic.setItem(i, 1,  _item(str(tanggal) if tanggal else ""))
            self.tabel_ic.setItem(i, 2,  _item(model or ""))
            self.tabel_ic.setItem(i, 3,  _item(op_no_st or ""))
            self.tabel_ic.setItem(i, 4,  _item(item or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_ic.setItem(i, 5,  _item(f"{float(qty):.0f}" if qty is not None else ""))
            self.tabel_ic.setItem(i, 6,  _item(penyebab or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_ic.setItem(i, 7,  _item(tindakan or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_ic.setItem(i, 8,  _item(faktor or ""))
            self.tabel_ic.setItem(i, 9,  _item(f"{float(stop_hr):.2f}" if stop_hr is not None else "0.00"))
            self.tabel_ic.setItem(i, 10, _item(f"{float(lost_hr):.2f}" if lost_hr is not None else "0.00"))
            it_status = _item(status or "")
            it_status.setForeground(QColor(220, 100, 100))
            self.tabel_ic.setItem(i, 11, it_status)
            self.tabel_ic.setRowHeight(i, 32)

        self.tabel_pp.setRowCount(0)
        self.lbl_info_pp.setText(f"— {len(inhouse_pending)} data")
        for i, row in enumerate(inhouse_pending):
            tanggal, model, op_no_st, item, qty, penyebab, tindakan, faktor, stop_hr, lost_hr = row
            self.tabel_pp.insertRow(i)
            self.tabel_pp.setItem(i, 0,  _item(i + 1))
            self.tabel_pp.setItem(i, 1,  _item(str(tanggal) if tanggal else ""))
            self.tabel_pp.setItem(i, 2,  _item(model or ""))
            self.tabel_pp.setItem(i, 3,  _item(op_no_st or ""))
            self.tabel_pp.setItem(i, 4,  _item(item or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_pp.setItem(i, 5,  _item(f"{float(qty):.0f}" if qty is not None else ""))
            self.tabel_pp.setItem(i, 6,  _item(penyebab or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_pp.setItem(i, 7,  _item(tindakan or "", Qt.AlignLeft | Qt.AlignVCenter))
            self.tabel_pp.setItem(i, 8,  _item(faktor or ""))
            self.tabel_pp.setItem(i, 9,  _item(f"{float(stop_hr):.2f}" if stop_hr is not None else "0.00"))
            self.tabel_pp.setItem(i, 10, _item(f"{float(lost_hr):.2f}" if lost_hr is not None else "0.00"))
            self.tabel_pp.setRowHeight(i, 32)

    def _export_ng_pending(self):
        ic_count = self.tabel_ic.rowCount()
        pp_count = self.tabel_pp.rowCount()
        if ic_count == 0 and pp_count == 0:
            QMessageBox.warning(self, "Peringatan", "Tidak ada data untuk diekspor.")
            return

        # Section name
        sec_text     = self.combo_section_ng.currentText()
        section_name = sec_text if sec_text != "Semua" else ""

        # Period string
        d_from = self.date_ng_from.date()
        d_to   = self.date_ng_to.date()
        _bln   = ["JAN", "FEB", "MAR", "APR", "MEI", "JUN",
                  "JUL", "AGU", "SEP", "OKT", "NOV", "DES"]
        if d_from.month() == d_to.month() and d_from.year() == d_to.year():
            period_str = f"{_bln[d_from.month() - 1]} {d_from.year()}"
        else:
            period_str = (
                f"{d_from.toString('dd/MM/yy')} - {d_to.toString('dd/MM/yy')}"
            )

        # Extract table data 
        def _read_table(tbl, ncols):
            """
            Baca ncols kolom dari QTableWidget.
            Sisipkan kolom Satuan (kosong) setelah Qty (index 5 → insert at 6).
            """
            rows = []
            for r in range(tbl.rowCount()):
                row = [
                    (tbl.item(r, c).text() if tbl.item(r, c) else "")
                    for c in range(ncols)
                ]
                row.insert(6, "")   # Satuan
                rows.append(row)
            return rows

        ng_rows      = _read_table(self.tabel_ic, 12)   # 12 cols → 13 setelah insert
        pending_rows = _read_table(self.tabel_pp, 11)   # 11 cols → 12 setelah insert

        try:
            filepath = export_inhouse_ng_pending(
                ng_rows, pending_rows, section_name, period_str
            )
            QMessageBox.information(self, "Export Berhasil", f"File disimpan di:\n{filepath}")
        except Exception as e:
            QMessageBox.critical(self, "Gagal Export", f"Gagal mengekspor: {e}")

    # Tab 4: Produktivitas Bulanan

    def _build_tab_produktivitas(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame(); card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f); fl.setContentsMargins(16, 12, 16, 12); fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_prod_section = QComboBox()
        self.combo_prod_section.setMinimumHeight(30)
        self.combo_prod_section.setMinimumWidth(155)
        self.combo_prod_section.setStyleSheet(_COMBO_STYLE)
        self.combo_prod_section.addItem("Semua", None)
        fl.addWidget(self.combo_prod_section)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_prod_bulan = QComboBox()
        self.combo_prod_bulan.setMinimumHeight(30)
        self.combo_prod_bulan.setMinimumWidth(115)
        self.combo_prod_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate(
            ["Januari","Februari","Maret","April","Mei","Juni",
             "Juli","Agustus","September","Oktober","November","Desember"], 1
        ):
            self.combo_prod_bulan.addItem(name, i)
        self.combo_prod_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_prod_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_prod_tahun = QLineEdit()
        self.input_prod_tahun.setMinimumHeight(30)
        self.input_prod_tahun.setFixedWidth(70)
        self.input_prod_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_prod_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_prod_tahun)

        fl.addStretch()
        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_produktivitas)
        fl.addWidget(btn_tampil)
        main.addWidget(card_f)

        # Content row
        content = QHBoxLayout(); content.setSpacing(12)

        # Left: breakdown table
        card_left = QFrame(); card_left.setStyleSheet(_CARD_STYLE)
        ll = QVBoxLayout(card_left); ll.setContentsMargins(16, 14, 16, 16); ll.setSpacing(8)
        lbl_bkdn = QLabel("Breakdown Jam Kerja")
        lbl_bkdn.setStyleSheet(
            "color: #212121; font-size: 12px; font-weight: bold;"
            " border-left: 3px solid #E60012; padding-left: 8px;"
        )
        ll.addWidget(lbl_bkdn)

        self.tabel_prod_bkdn = QTableWidget()
        self.tabel_prod_bkdn.setColumnCount(2)
        self.tabel_prod_bkdn.setHorizontalHeaderLabels(["Kategori", "Jam (H)"])
        self.tabel_prod_bkdn.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tabel_prod_bkdn.horizontalHeader().setSectionResizeMode(1, QHeaderView.Fixed)
        self.tabel_prod_bkdn.setColumnWidth(1, 90)
        self.tabel_prod_bkdn.verticalHeader().setVisible(False)
        self.tabel_prod_bkdn.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_prod_bkdn.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_prod_bkdn.setMinimumHeight(300)
        self.tabel_prod_bkdn.setStyleSheet(_TABLE_STYLE)
        ll.addWidget(self.tabel_prod_bkdn)
        content.addWidget(card_left, 3)

        # Right: ratio + summary
        card_right = QFrame(); card_right.setStyleSheet(_CARD_STYLE)
        rl = QVBoxLayout(card_right); rl.setContentsMargins(16, 14, 16, 16); rl.setSpacing(12)

        lbl_rt = QLabel("Ratio Produktivitas")
        lbl_rt.setStyleSheet("color: #6B7280; font-size: 11px; font-weight: bold;")
        lbl_rt.setAlignment(Qt.AlignCenter)
        rl.addWidget(lbl_rt)

        ratio_frame = QFrame()
        ratio_frame.setObjectName("ratioFrame")
        ratio_frame.setStyleSheet(
            "#ratioFrame { background-color: #F3F4F6; border: 2px solid #D0D0D0; }"
        )
        rfl = QVBoxLayout(ratio_frame); rfl.setContentsMargins(16, 24, 16, 24)
        self._lbl_ratio_big = QLabel("—")
        self._lbl_ratio_big.setStyleSheet(
            "color: #27ae60; font-size: 40px; font-weight: bold;"
        )
        self._lbl_ratio_big.setAlignment(Qt.AlignCenter)
        rfl.addWidget(self._lbl_ratio_big)
        rl.addWidget(ratio_frame)

        sum_frame = QFrame()
        sum_frame.setStyleSheet(
            "QFrame { background-color: #F8F9FA; border: 1px solid #D1D5DB; }"
        )
        sg = QGridLayout(sum_frame); sg.setContentsMargins(12, 10, 12, 10); sg.setSpacing(6)
        sg.setColumnStretch(1, 1)

        def _srow(row, label, attr):
            l = QLabel(label); l.setStyleSheet("color: #6B7280; font-size: 11px;")
            v = QLabel("—")
            v.setStyleSheet("color: #212121; font-size: 11px; font-weight: bold;")
            v.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            sg.addWidget(l, row, 0); sg.addWidget(v, row, 1)
            setattr(self, attr, v)

        _srow(0, "Total Hour",   "_lbl_prod_total")
        _srow(1, "Loss Time",    "_lbl_prod_bal")
        _srow(2, "Laporan",      "_lbl_prod_count")
        rl.addWidget(sum_frame)
        rl.addStretch()
        content.addWidget(card_right, 2)

        main.addLayout(content)
        main.addStretch()
        scroll.setWidget(container)
        outer.addWidget(scroll)
        return tab

    def load_produktivitas(self):
        section_id = self.combo_prod_section.currentData()
        bulan = self.combo_prod_bulan.currentData()
        try:
            tahun = int(self.input_prod_tahun.text().strip())
        except ValueError:
            QMessageBox.warning(self, "Peringatan", "Tahun tidak valid.")
            return

        try:
            data = get_monthly_productivity(section_id, bulan, tahun)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data: {e}")
            return

        tbl = self.tabel_prod_bkdn
        tbl.setRowCount(0)

        loss_hour  = sum(c["hours"] for c in data["categories"])
        total_hour = data["total_hour"]

        def _grp_row(text):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it = QTableWidgetItem(text)
            it.setFlags(Qt.ItemIsEnabled)
            it.setBackground(QColor("#F8F9FA"))
            it.setForeground(QColor("#E60012"))
            f = it.font(); f.setBold(True); it.setFont(f)
            tbl.setItem(r, 0, it)
            it2 = QTableWidgetItem("")
            it2.setFlags(Qt.ItemIsEnabled)
            it2.setBackground(QColor("#F8F9FA"))
            tbl.setItem(r, 1, it2)
            tbl.setRowHeight(r, 26)

        def _subgrp_row(text):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it = QTableWidgetItem(f"  {text}")
            it.setFlags(Qt.ItemIsEnabled)
            it.setForeground(QColor("#606060"))
            f = it.font(); f.setItalic(True); it.setFont(f)
            tbl.setItem(r, 0, it)
            it2 = QTableWidgetItem("")
            it2.setFlags(Qt.ItemIsEnabled)
            tbl.setItem(r, 1, it2)
            tbl.setRowHeight(r, 20)

        def _data_row(name, hours, fg=None):
            r = tbl.rowCount()
            tbl.insertRow(r)
            it_n = QTableWidgetItem(f"    {name}")
            it_n.setFlags(Qt.ItemIsEnabled)
            it_h = QTableWidgetItem(f"{hours:.4f}")
            it_h.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            it_h.setFlags(Qt.ItemIsEnabled)
            if fg:
                it_n.setForeground(QColor(fg))
                it_h.setForeground(QColor(fg))
            tbl.setItem(r, 0, it_n)
            tbl.setItem(r, 1, it_h)
            tbl.setRowHeight(r, 28)

        # Produktif 
        _grp_row("PRODUKTIF")
        _data_row("Process", data["process_hour"], "#80c880")

        # Non-Produktif (line stop by category/group)

        groups: dict[str, list] = {}
        for cat in data["categories"]:
            groups.setdefault(cat["group"], []).append(cat)

        if groups:
            _grp_row("NON-PRODUKTIF (Line Stop)")
            for grp_name, cats in groups.items():
                _subgrp_row(grp_name)
                for cat in cats:
                    _data_row(cat["name"], cat["hours"])

        # Waktu Tetap 
        _grp_row("WAKTU TETAP")
        _data_row("Preparation", data["prep_hour"])
        _data_row("Sholat",      data["sholat_hour"])

        # Ketidakhadiran
        _grp_row("KETIDAKHADIRAN")
        _data_row("Absence", data["absence_hour"])

        # Divider + TOTAL
        r = tbl.rowCount()
        tbl.insertRow(r)
        for c in range(2):
            it = QTableWidgetItem()
            it.setFlags(Qt.ItemIsEnabled)
            it.setBackground(QColor("#404040"))
            tbl.setItem(r, c, it)
        tbl.setRowHeight(r, 2)

        r = tbl.rowCount()
        tbl.insertRow(r)
        it_n = QTableWidgetItem("TOTAL")
        it_n.setFlags(Qt.ItemIsEnabled)
        fn = it_n.font(); fn.setBold(True); it_n.setFont(fn)
        it_n.setForeground(QColor("#212121"))
        it_h = QTableWidgetItem(f"{total_hour:.4f}")
        it_h.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
        it_h.setFlags(Qt.ItemIsEnabled)
        fh = it_h.font(); fh.setBold(True); it_h.setFont(fh)
        it_h.setForeground(QColor("#212121"))
        tbl.setItem(r, 0, it_n)
        tbl.setItem(r, 1, it_h)
        tbl.setRowHeight(r, 30)

        # Summary
        balance = loss_hour

        ratio = (data["process_hour"] / total_hour * 100) if total_hour > 0 else 0.0
        ratio_color = (
            "#27ae60" if ratio >= 80
            else "#f39c12" if ratio >= 60
            else "#E60012"
        )
        self._lbl_ratio_big.setText(f"{ratio:.1f}%")
        self._lbl_ratio_big.setStyleSheet(
            f"color: {ratio_color}; font-size: 40px; font-weight: bold;"
        )
        self._lbl_prod_total.setText(f"{total_hour:.2f} H")

        bal_color = "#27ae60" if balance < 0.01 else "#E60012"
        self._lbl_prod_bal.setText(f"{balance:.4f}")
        self._lbl_prod_bal.setStyleSheet(
            f"color: {bal_color}; font-size: 11px; font-weight: bold;"
        )
        self._lbl_prod_count.setText(f"{data['report_count']} laporan")

    @staticmethod
    def _flabel(text):
        lbl = QLabel(text)
        lbl.setStyleSheet("color: #6B7280; font-size: 11px;")
        return lbl

    # Tab 5: Volume Produksi

    def _build_tab_volume(self):
        tab = QWidget()
        tab.setStyleSheet("background: transparent;")
        outer = QVBoxLayout(tab)
        outer.setContentsMargins(0, 0, 0, 0)

        scroll_v = QScrollArea()
        scroll_v.setWidgetResizable(True)
        scroll_v.setStyleSheet("QScrollArea { border: none; background: transparent; }")
        container = QWidget()
        container.setStyleSheet("background: transparent;")
        main = QVBoxLayout(container)
        main.setContentsMargins(15, 15, 15, 15)
        main.setSpacing(12)

        # Filter card
        card_f = QFrame()
        card_f.setStyleSheet(_CARD_STYLE)
        fl = QHBoxLayout(card_f)
        fl.setContentsMargins(16, 12, 16, 12)
        fl.setSpacing(10)

        fl.addWidget(self._flabel("Shop"))
        self.combo_vol_section = QComboBox()
        self.combo_vol_section.setMinimumHeight(30)
        self.combo_vol_section.setMinimumWidth(155)
        self.combo_vol_section.setStyleSheet(_COMBO_STYLE)
        fl.addWidget(self.combo_vol_section)

        fl.addWidget(self._flabel("Bulan"))
        self.combo_vol_bulan = QComboBox()
        self.combo_vol_bulan.setMinimumHeight(30)
        self.combo_vol_bulan.setMinimumWidth(115)
        self.combo_vol_bulan.setStyleSheet(_COMBO_STYLE)
        for i, name in enumerate([
            "Januari", "Februari", "Maret", "April", "Mei", "Juni",
            "Juli", "Agustus", "September", "Oktober", "November", "Desember",
        ], 1):
            self.combo_vol_bulan.addItem(name, i)
        self.combo_vol_bulan.setCurrentIndex(QDate.currentDate().month() - 1)
        fl.addWidget(self.combo_vol_bulan)

        fl.addWidget(self._flabel("Tahun"))
        self.input_vol_tahun = QLineEdit()
        self.input_vol_tahun.setMinimumHeight(30)
        self.input_vol_tahun.setFixedWidth(70)
        self.input_vol_tahun.setStyleSheet(_INPUT_STYLE)
        self.input_vol_tahun.setText(str(QDate.currentDate().year()))
        fl.addWidget(self.input_vol_tahun)

        fl.addStretch()
        btn_tampil = QPushButton("Tampilkan")
        btn_tampil.setMinimumSize(90, 32)
        btn_tampil.setStyleSheet(_BTN_CARI)
        btn_tampil.clicked.connect(self.load_volume_produksi)
        fl.addWidget(btn_tampil)
        main.addWidget(card_f)

        # Table card with horizontal scroll
        card_t = QFrame()
        card_t.setStyleSheet(_CARD_STYLE)
        tl = QVBoxLayout(card_t)
        tl.setContentsMargins(16, 16, 16, 16)
        tl.setSpacing(8)

        self.lbl_info_vol = QLabel("Pilih Shop dan klik Tampilkan.")
        self.lbl_info_vol.setStyleSheet("color: #6B7280; font-size: 10px;")
        tl.addWidget(self.lbl_info_vol)

        scroll_h = QScrollArea()
        scroll_h.setWidgetResizable(True)
        scroll_h.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_h.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll_h.setStyleSheet("QScrollArea { border: none; background: transparent; }")

        self.tabel_volume = QTableWidget()
        self.tabel_volume.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.tabel_volume.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.tabel_volume.verticalHeader().setVisible(False)
        self.tabel_volume.setStyleSheet(_TABLE_STYLE)
        self.tabel_volume.horizontalHeader().setDefaultSectionSize(28)
        self.tabel_volume.horizontalHeader().setMinimumSectionSize(20)
        self.tabel_volume.setMinimumHeight(300)

        scroll_h.setWidget(self.tabel_volume)
        tl.addWidget(scroll_h)
        main.addWidget(card_t)
        main.addStretch()
        scroll_v.setWidget(container)
        outer.addWidget(scroll_v)
        return tab

    def load_volume_produksi(self):
        section_id = self.combo_vol_section.currentData()
        if section_id is None:
            QMessageBox.warning(self, "Peringatan", "Pilih Shop terlebih dahulu.")
            return
        bulan = self.combo_vol_bulan.currentData()
        try:
            tahun = int(self.input_vol_tahun.text().strip())
        except ValueError:
            return

        try:
            result = get_production_volume(section_id, bulan, tahun)
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Gagal memuat data: {e}")
            return

        tbl  = self.tabel_volume
        days = result["days"]

        tbl.setColumnCount(2 + days + 1)
        headers = ["Model", "Shift"] + [str(d) for d in range(1, days + 1)] + ["TOTAL"]
        tbl.setHorizontalHeaderLabels(headers)

        hh = tbl.horizontalHeader()
        hh.setSectionResizeMode(QHeaderView.Stretch)
        hh.setSectionResizeMode(0, QHeaderView.Fixed)
        hh.setSectionResizeMode(1, QHeaderView.Fixed)
        hh.setSectionResizeMode(2 + days, QHeaderView.Fixed)

        tbl.setColumnWidth(0, 100)
        tbl.setColumnWidth(1, 90)
        tbl.setColumnWidth(2 + days, 70)

        tbl.setRowCount(0)

        row_idx = 0
        for model in result["models"]:
            for shift in result["shifts"]:
                tbl.insertRow(row_idx)
                tbl.setRowHeight(row_idx, 26)

                tbl.setItem(row_idx, 0, _vol_item(model, bold=(shift == result["shifts"][0])))
                tbl.setItem(row_idx, 1, _vol_item(shift, center=True))

                key      = (model, shift)
                day_data = result["data"].get(key, {})
                for d in range(1, days + 1):
                    val = day_data.get(d)
                    it  = QTableWidgetItem(str(val) if val else "")
                    it.setTextAlignment(Qt.AlignCenter)
                    it.setFlags(Qt.ItemIsEnabled)
                    if val:
                        it.setBackground(QColor("#FFFFFF"))
                    else:
                        it.setBackground(QColor("#F9FAFB"))
                        it.setForeground(QColor("#D1D5DB"))
                    tbl.setItem(row_idx, d + 1, it)

                total    = result["totals"].get(key, 0)
                it_total = QTableWidgetItem(str(total) if total else "")
                it_total.setTextAlignment(Qt.AlignCenter)
                it_total.setFlags(Qt.ItemIsEnabled)
                it_total.setForeground(QColor("#E60012"))
                fn = it_total.font(); fn.setBold(True); it_total.setFont(fn)
                tbl.setItem(row_idx, 2 + days, it_total)

                row_idx += 1

        bulan_nama = self.combo_vol_bulan.currentText()
        self.lbl_info_vol.setText(
            f"{row_idx} baris — {bulan_nama} {tahun} | {self.combo_vol_section.currentText()}"
        )


# Module-level helpers

def _vol_item(text: str, bold: bool = False, center: bool = False) -> QTableWidgetItem:
    it = QTableWidgetItem(text)
    it.setTextAlignment(Qt.AlignCenter if center else Qt.AlignLeft | Qt.AlignVCenter)
    it.setFlags(Qt.ItemIsEnabled)
    if bold:
        f = it.font(); f.setBold(True); it.setFont(f)
    return it
